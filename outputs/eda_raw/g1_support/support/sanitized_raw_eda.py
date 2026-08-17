"""Privacy-safe, aggregate-only raw EDA for the G1 support branch.

This module never persists raw text, raw previews, pair identifiers, URLs, or
text hashes. Raw KO/EN strings are used only transiently to update aggregate
length/noise counters and salted in-memory duplicate keys.
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook


HTML_RE = re.compile(r"<[A-Za-z/!][^>\n]{0,200}>")
URL_RE = re.compile(r"(?i)\b(?:https?://|www\.)\S+")
EMAIL_RE = re.compile(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?\d[\d ()-]{7,}\d)(?!\d)")
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
ZERO_WIDTH_RE = re.compile(r"[\u200b\u200c\u200d\u2060\ufeff]")
HANGUL_RE = re.compile(r"[가-힣ㄱ-ㅎㅏ-ㅣ]")
LATIN_RE = re.compile(r"[A-Za-z]")
SAFE_CATEGORY_FIELDS = {
    "data_set", "domain", "subdomain", "source", "style",
    "source_language", "target_language", "license",
    "대분류", "소분류", "상황", "자동분류1", "자동분류2", "자동분류3",
    "언론사", "키워드", "지자체",
}
RAW_COLUMN_NAMES = {"ko", "en", "ko_preview", "en_preview", "원문", "번역문", "URL"}


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def configure_plotting() -> None:
    available = {font.name for font in __import__("matplotlib").font_manager.fontManager.ttflist}
    font = "NanumGothic" if "NanumGothic" in available else "DejaVu Sans"
    plt.rcParams.update({
        "font.family": font,
        "axes.unicode_minus": False,
        "figure.facecolor": "#F8FAFC",
        "axes.facecolor": "white",
        "axes.titleweight": "bold",
        "axes.grid": True,
        "grid.alpha": 0.18,
        "savefig.dpi": 160,
    })


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def iter_json_data(path: Path) -> Iterator[dict[str, Any]]:
    decoder = json.JSONDecoder()
    with path.open("r", encoding="utf-8-sig", errors="strict") as handle:
        prefix = ""
        match = None
        while match is None:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                raise ValueError(f"data[] not found: {path.name}")
            prefix += chunk
            match = re.search(r'"data"\s*:\s*\[', prefix)
        buffer, position = prefix[match.end():], 0
        while True:
            while True:
                while position < len(buffer) and buffer[position] in " \t\r\n,":
                    position += 1
                if position < len(buffer):
                    break
                buffer, position = handle.read(4 * 1024 * 1024), 0
                if not buffer:
                    raise ValueError(f"unexpected EOF: {path.name}")
            if buffer[position] == "]":
                return
            while True:
                try:
                    value, end = decoder.raw_decode(buffer, position)
                    break
                except json.JSONDecodeError:
                    remainder = buffer[position:]
                    chunk = handle.read(4 * 1024 * 1024)
                    if not chunk:
                        raise
                    buffer, position = remainder + chunk, 0
            if not isinstance(value, dict):
                raise TypeError("data[] member is not an object")
            yield value
            position = end
            if position > 8 * 1024 * 1024:
                buffer, position = buffer[position:], 0


def iter_xlsx_rows(path: Path) -> tuple[list[str], Iterator[dict[str, Any]]]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    raw_header = next(iterator, ())
    last = max((i for i, value in enumerate(raw_header) if value not in (None, "")), default=-1)
    header = [str(value) if value not in (None, "") else f"__unnamed_{i+1}" for i, value in enumerate(raw_header[:last + 1])]

    def rows() -> Iterator[dict[str, Any]]:
        try:
            for values in iterator:
                values = values[:len(header)]
                if not any(value not in (None, "") for value in values):
                    continue
                yield {header[i]: value for i, value in enumerate(values)}
        finally:
            workbook.close()

    return header, rows()


def json_path_meta(relative: str) -> tuple[str, str]:
    split = "TRAIN" if "/1.Training/" in f"/{relative}" else "VALID" if "/2.Validation/" in f"/{relative}" else "UNSPLIT"
    name = Path(relative).name
    direction = "EN_TO_KO" if "영한" in name else "KO_TO_EN" if "한영" in name else "UNKNOWN"
    return split, direction


def xlsx_family(path: Path) -> str:
    name = path.name
    if name.startswith("1_구어체"):
        return "구어체"
    if name.startswith("2_대화체"):
        return "대화체"
    if name.startswith("3_문어체_뉴스"):
        return "문어체뉴스"
    if name.startswith("4_문어체_한국문화"):
        return "한국문화"
    if name.startswith("5_문어체_조례"):
        return "조례"
    if name.startswith("6_문어체_지자체"):
        return "지자체웹"
    return "기타"


def scalar_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def histogram_summary(counter: Counter[int]) -> dict[str, float | int]:
    total = sum(counter.values())
    if not total:
        return {"count": 0, "min": 0, "p50": 0, "p90": 0, "p95": 0, "p99": 0, "max": 0, "mean": 0.0}
    ordered = sorted(counter.items())

    def quantile(probability: float) -> int:
        target, cumulative = max(1, math.ceil(total * probability)), 0
        for value, count in ordered:
            cumulative += count
            if cumulative >= target:
                return value
        return ordered[-1][0]

    return {
        "count": total,
        "min": ordered[0][0],
        "p50": quantile(.50),
        "p90": quantile(.90),
        "p95": quantile(.95),
        "p99": quantile(.99),
        "max": ordered[-1][0],
        "mean": sum(value * count for value, count in ordered) / total,
    }


def _text_flags(text: str, language: str) -> dict[str, bool]:
    return {
        "empty": text.strip() == "",
        "leading_or_trailing_whitespace": text != text.strip(),
        "repeated_whitespace": bool(re.search(r"[ \t]{2,}", text)),
        "newline_or_tab": "\n" in text or "\r" in text or "\t" in text,
        "html_like": bool(HTML_RE.search(text)),
        "url_like": bool(URL_RE.search(text)),
        "email_like": bool(EMAIL_RE.search(text)),
        "phone_like": bool(PHONE_RE.search(text)),
        "control_character": bool(CONTROL_RE.search(text)),
        "zero_width": bool(ZERO_WIDTH_RE.search(text)),
        "replacement_character": "\ufffd" in text,
        "expected_script_absent": bool(text) and (not HANGUL_RE.search(text) if language == "KO" else not LATIN_RE.search(text)),
        "nfkc_change": unicodedata.normalize("NFKC", text) != text,
    }


def _pair_digest(ko: str, en: str, salt: bytes) -> bytes:
    digest = hashlib.blake2b(key=salt, digest_size=16)
    for text in (ko, en):
        encoded = text.encode("utf-8", errors="strict")
        digest.update(len(encoded).to_bytes(8, "big"))
        digest.update(encoded)
    return digest.digest()


def _select_data_files(raw_root: Path, kind: str) -> list[Path]:
    if kind == "json":
        return sorted(path for path in raw_root.rglob("*.json") if "/라벨링데이터/" in f"/{path.relative_to(raw_root).as_posix()}")
    return sorted(raw_root.rglob("*.xlsx"))


def profile_dataset(config: dict[str, Any]) -> dict[str, Any]:
    raw_root = Path(config["raw_root"])
    if not raw_root.is_dir():
        raise FileNotFoundError(raw_root)
    started = now_iso()
    all_files = sorted(path for path in raw_root.rglob("*") if path.is_file())
    start_state = {path: (path.stat().st_size, path.stat().st_mtime_ns) for path in all_files}
    inventory_rows = []
    for path in all_files:
        relative = path.relative_to(raw_root).as_posix()
        split, direction = json_path_meta(relative) if path.suffix.lower() == ".json" else ("UNSPLIT", "FIELD_ORDER_ONLY")
        inventory_rows.append({
            "relative_path": relative,
            "bytes": path.stat().st_size,
            "mtime_ns": path.stat().st_mtime_ns,
            "sha256": sha256_file(path),
            "format": path.suffix.lower().lstrip(".").upper(),
            "split": split,
            "direction": direction,
            "stability_state": "STABLE_FILE_PENDING_END_CHECK",
        })

    group_counts: Counter[str] = Counter()
    schema: dict[tuple[str, str], dict[str, Any]] = {}
    length_hists: dict[tuple[str, str, str], Counter[int]] = defaultdict(Counter)
    noise: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    categories: Counter[tuple[str, str, str]] = Counter()
    duplicate_rows = 0
    duplicate_groups: set[bytes] = set()
    seen_pairs: set[bytes] = set()
    salt = os.urandom(32)

    data_files = _select_data_files(raw_root, config["kind"])
    for path in data_files:
        relative = path.relative_to(raw_root).as_posix()
        if config["kind"] == "json":
            split, direction = json_path_meta(relative)
            group = f"{split} · {direction}"
            iterator = iter_json_data(path)
            ko_field, en_field = "ko", "en"
        else:
            _, iterator = iter_xlsx_rows(path)
            group = xlsx_family(path)
            ko_field, en_field = "원문", "번역문"
        for row in iterator:
            prior = group_counts[group]
            group_counts[group] += 1
            for field in row:
                schema.setdefault((group, field), {"absent": prior, "null": 0, "empty": 0, "types": Counter()})
            for (schema_group, field), state in schema.items():
                if schema_group != group:
                    continue
                if field not in row:
                    state["absent"] += 1
                else:
                    value = row[field]
                    state["types"][scalar_type(value)] += 1
                    if value is None:
                        state["null"] += 1
                    elif isinstance(value, str) and value.strip() == "":
                        state["empty"] += 1
            ko, en = str(row.get(ko_field) or ""), str(row.get(en_field) or "")
            for language, text in (("KO", ko), ("EN", en)):
                length_hists[(group, language, "codepoints")][len(text)] += 1
                length_hists[(group, language, "utf8_bytes")][len(text.encode("utf-8"))] += 1
                for flag, present in _text_flags(text, language).items():
                    if present:
                        noise[(group, language)][flag] += 1
            pair_hash = _pair_digest(ko, en, salt)
            if pair_hash in seen_pairs:
                duplicate_rows += 1
                duplicate_groups.add(pair_hash)
            else:
                seen_pairs.add(pair_hash)
            for field in SAFE_CATEGORY_FIELDS.intersection(row):
                value = row.get(field)
                if value not in (None, ""):
                    rendered = str(value)
                    if len(rendered) <= 120 and not URL_RE.search(rendered) and not EMAIL_RE.search(rendered) and not PHONE_RE.search(rendered):
                        categories[(group, field, rendered)] += 1

    end_state = {path: (path.stat().st_size, path.stat().st_mtime_ns) for path in all_files}
    changed = [path.relative_to(raw_root).as_posix() for path in all_files if start_state[path] != end_state[path]]
    if changed:
        raise RuntimeError(f"DOWNLOAD_IN_PROGRESS: {len(changed)} changed files")

    inventory = pd.DataFrame(inventory_rows)
    inventory["stability_state"] = "STABLE_FILE"
    record_counts = pd.DataFrame([{"group": group, "records": count} for group, count in sorted(group_counts.items())])
    schema_rows = []
    for (group, field), state in sorted(schema.items()):
        total = group_counts[group]
        schema_rows.append({
            "group": group,
            "field": field,
            "records": total,
            "absent": state["absent"],
            "null": state["null"],
            "empty": state["empty"],
            "missing_or_empty_rate": (state["absent"] + state["null"] + state["empty"]) / max(1, total),
            "type_counts": json.dumps(dict(state["types"]), sort_keys=True),
            "raw_value_exported": False,
        })
    length_rows = []
    for (group, language, metric), counter in sorted(length_hists.items()):
        length_rows.append({"group": group, "language": language, "metric": metric, **histogram_summary(counter)})
    noise_rows = []
    for (group, language), counter in sorted(noise.items()):
        total = group_counts[group]
        for metric, count in sorted(counter.items()):
            noise_rows.append({
                "group": group,
                "language": language,
                "metric": metric,
                "row_count": count,
                "records": total,
                "rate": count / max(1, total),
                "classification": "SOFT_OBSERVATION_ONLY",
            })
    category_rows = [
        {"group": group, "dimension": dimension, "category": value, "record_count": count}
        for (group, dimension, value), count in categories.items()
    ]
    duplicates = pd.DataFrame([{
        "scope": "local_dataset",
        "records": int(sum(group_counts.values())),
        "duplicate_pair_rows_after_first": duplicate_rows,
        "distinct_duplicate_pair_groups": len(duplicate_groups),
        "method": "salted in-memory BLAKE2b-128 exact raw KO+EN; no hashes persisted",
        "classification": "DUPLICATE_CANDIDATE_ONLY",
    }])
    summary = {
        "dataset_local_id": config["dataset_local_id"],
        "audit_timestamp": started,
        "audit_completed_timestamp": now_iso(),
        "physical_file_count": len(all_files),
        "profiled_data_file_count": len(data_files),
        "logical_record_count": int(sum(group_counts.values())),
        "population_scope": "FULL_POPULATION_AGGREGATES",
        "raw_text_exported": False,
        "text_hash_exported": False,
        "status": "G1_SANITIZED_SUPPORT / NOT QC ACCEPTANCE",
    }
    return {
        "config": config,
        "summary": summary,
        "inventory": inventory,
        "record_counts": record_counts,
        "schema": pd.DataFrame(schema_rows),
        "length_summary": pd.DataFrame(length_rows),
        "noise": pd.DataFrame(noise_rows),
        "categories": pd.DataFrame(category_rows),
        "duplicates": duplicates,
        "length_hists": length_hists,
    }


def save_profile(result: dict[str, Any], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = result["config"]["artifact_prefix"]
    paths = []
    for key in ["inventory", "record_counts", "schema", "length_summary", "noise", "categories", "duplicates"]:
        path = output_dir / f"{prefix}_{key}.csv"
        result[key].to_csv(path, index=False, encoding="utf-8")
        paths.append(path)
    summary_path = output_dir / f"{prefix}_profile_summary.json"
    summary_path.write_text(json.dumps(result["summary"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    paths.append(summary_path)
    return paths


def _save_figure(fig: plt.Figure, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, bbox_inches="tight")
    plt.show()
    return path


def plot_profile(result: dict[str, Any], figure_dir: Path) -> list[Path]:
    configure_plotting()
    prefix = result["config"]["artifact_prefix"]
    made = []
    counts = result["record_counts"].sort_values("records")
    fig, axes = plt.subplots(1, 2, figsize=(15, 5), constrained_layout=True)
    axes[0].barh(counts.group, counts.records, color="#0F766E")
    axes[0].xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{x/1e6:.1f}M" if x >= 1e6 else f"{x/1e3:.0f}k"))
    axes[0].set_title("Full-population logical record counts")
    axes[0].set_xlabel("records")
    inv = result["inventory"].copy()
    inv["MiB"] = inv.bytes / 2**20
    inv["file"] = inv.relative_path.map(lambda value: Path(value).name[:36])
    axes[1].barh(inv.file, inv.MiB, color="#475569")
    axes[1].set_title("Physical raw files")
    axes[1].set_xlabel("MiB")
    fig.suptitle(f"{result['config']['dataset_local_id']} — sanitized structure/scale", fontsize=16)
    made.append(_save_figure(fig, figure_dir / f"{prefix}_01_structure.png"))

    fig, axes = plt.subplots(1, 2, figsize=(15, 6), constrained_layout=True)
    for axis, metric in zip(axes, ["codepoints", "utf8_bytes"]):
        for group in counts.group:
            for language, linestyle in [("KO", "-"), ("EN", "--")]:
                counter = result["length_hists"].get((group, language, metric), Counter())
                summary = histogram_summary(counter)
                limit = int(summary["p99"])
                xs = np.arange(limit + 1)
                cumulative = np.cumsum([counter.get(int(x), 0) for x in xs]) / max(1, sum(counter.values()))
                axis.plot(xs, cumulative, linestyle=linestyle, linewidth=1.5, label=f"{group} · {language}")
        axis.set_title(f"Raw {metric} ECDF (≤ group p99)")
        axis.set_xlabel(metric)
        axis.set_ylabel("cumulative share")
        axis.legend(fontsize=7)
    fig.suptitle(f"{result['config']['dataset_local_id']} — aggregate KO/EN lengths", fontsize=16)
    made.append(_save_figure(fig, figure_dir / f"{prefix}_02_lengths.png"))

    categories = result["categories"]
    dimensions = categories.groupby("dimension").record_count.sum().nlargest(4).index.tolist() if len(categories) else []
    fig, axes = plt.subplots(2, 2, figsize=(15, 10), constrained_layout=True)
    for axis, dimension in zip(axes.flat, dimensions):
        top = categories[categories.dimension.eq(dimension)].groupby("category").record_count.sum().nlargest(12).sort_values()
        axis.barh([str(value)[:42] for value in top.index], top.values, color="#7C3AED")
        axis.set_title(f"{dimension}: top aggregate categories")
        axis.set_xlabel("records")
    for axis in list(axes.flat)[len(dimensions):]:
        axis.axis("off")
    fig.suptitle(f"{result['config']['dataset_local_id']} — category composition", fontsize=16)
    made.append(_save_figure(fig, figure_dir / f"{prefix}_03_categories.png"))

    noise = result["noise"].copy()
    noise["label"] = noise.language + " · " + noise.metric
    top = noise.groupby("label", as_index=False)[["row_count", "records"]].sum()
    top["rate"] = top.row_count / top.records
    top = top.nlargest(16, "rate").sort_values("rate")
    fig, ax = plt.subplots(figsize=(12, 7), constrained_layout=True)
    ax.barh(top.label, top.rate, color="#D97706")
    ax.set_title("Soft noise/preprocessing indicators — no automatic exclusion")
    ax.set_xlabel("row rate")
    fig.suptitle(f"{result['config']['dataset_local_id']} — aggregate quality signals", fontsize=16)
    made.append(_save_figure(fig, figure_dir / f"{prefix}_04_noise.png"))
    return made


def _save_frames(frames: dict[str, pd.DataFrame], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for name, frame in frames.items():
        path = output_dir / f"{name}.csv"
        frame.to_csv(path, index=False, encoding="utf-8")
        paths.append(path)
    return paths


def targeted_025(raw_root: Path, output_dir: Path, figure_dir: Path) -> dict[str, pd.DataFrame]:
    direction_split_domain = Counter()
    direction_source_domain = Counter()
    mapping = Counter()
    packed_counts: dict[bytes, int] = {}
    salt = os.urandom(32)
    direction_shift = {"EN_TO_KO": 0, "KO_TO_EN": 16}
    split_shift = {"TRAIN": 32, "VALID": 48}
    total = 0
    for path in _select_data_files(raw_root, "json"):
        split, direction = json_path_meta(path.relative_to(raw_root).as_posix())
        for row in iter_json_data(path):
            domain = str(row.get("domain") or "<MISSING>")
            source = str(row.get("source") or "<MISSING>")
            direction_split_domain[(direction, split, domain)] += 1
            direction_source_domain[(direction, source, domain)] += 1
            mapped = {"일상생활": "general", "해외고객과의채팅": "dialogue", "해외영업": "other"}.get(domain, "UNMAPPED_PREVIEW")
            mapping[(domain, mapped)] += 1
            key = _pair_digest(str(row.get("ko") or ""), str(row.get("en") or ""), salt)
            packed_counts[key] = packed_counts.get(key, 0) + (1 << direction_shift[direction]) + (1 << split_shift[split])
            total += 1

    def unpack(value: int) -> tuple[int, int, int, int]:
        return tuple((value >> shift) & 0xFFFF for shift in (0, 16, 32, 48))

    multiplicity = Counter()
    within_enko = within_koen = within_train = within_valid = 0
    mirror_rows = mirror_groups = split_rows = split_groups = duplicate_rows = 0
    for packed in packed_counts.values():
        enko, koen, train, valid = unpack(packed)
        multiplicity[enko + koen] += 1
        duplicate_rows += max(enko + koen - 1, 0)
        within_enko += max(enko - 1, 0)
        within_koen += max(koen - 1, 0)
        within_train += max(train - 1, 0)
        within_valid += max(valid - 1, 0)
        mirror_rows += min(enko, koen)
        mirror_groups += int(enko > 0 and koen > 0)
        split_rows += min(train, valid)
        split_groups += int(train > 0 and valid > 0)

    dsd = pd.DataFrame([
        {"direction": d, "split": s, "raw_domain": domain, "record_count": count}
        for (d, s, domain), count in direction_split_domain.items()
    ]).sort_values(["direction", "split", "raw_domain"])
    dsource = pd.DataFrame([
        {"direction": d, "raw_source": source, "raw_domain": domain, "record_count": count}
        for (d, source, domain), count in direction_source_domain.items()
    ]).sort_values(["direction", "raw_source", "raw_domain"])
    mechanisms = pd.DataFrame([
        {"mechanism": "within EN_TO_KO", "candidate_row_count": within_enko},
        {"mechanism": "within KO_TO_EN", "candidate_row_count": within_koen},
        {"mechanism": "EN_TO_KO <-> KO_TO_EN", "candidate_row_count": mirror_rows},
        {"mechanism": "within TRAIN", "candidate_row_count": within_train},
        {"mechanism": "within VALID", "candidate_row_count": within_valid},
        {"mechanism": "TRAIN <-> VALID", "candidate_row_count": split_rows},
    ])
    explanation = pd.DataFrame([{
        "full_population_rows": total,
        "distinct_exact_pair_groups": len(packed_counts),
        "duplicate_rows_after_first": duplicate_rows,
        "direction_mirror_matchable_rows": mirror_rows,
        "direction_mirror_group_count": mirror_groups,
        "direction_mirror_share_of_duplicate_rows": mirror_rows / max(1, duplicate_rows),
        "cross_split_matchable_rows": split_rows,
        "cross_split_group_count": split_groups,
        "classification": "STRUCTURAL_CANDIDATE_ONLY",
    }])
    mult = pd.DataFrame([
        {"duplicate_group_size": size, "exact_pair_group_count": count, "rows_in_groups": size * count}
        for size, count in sorted(multiplicity.items()) if size >= 2
    ])
    sbs = dsource[dsource.raw_source.eq("SBS")]
    sbs_check = pd.DataFrame([{
        "SBS_record_count": int(sbs.record_count.sum()),
        "SBS_outside_KO_TO_EN_count": int(sbs.loc[~sbs.direction.eq("KO_TO_EN"), "record_count"].sum()),
        "SBS_outside_daily_life_domain_count": int(sbs.loc[~sbs.raw_domain.eq("일상생활"), "record_count"].sum()),
        "observed_only_in_KO_TO_EN_and_daily_life": bool(len(sbs) and sbs.direction.eq("KO_TO_EN").all() and sbs.raw_domain.eq("일상생활").all()),
    }])
    crowd = dsource[dsource.raw_source.isin(["크라우드 소싱", "크라우드소싱"])].copy()
    crowd["classification"] = "DISTINCT_RAW_CATEGORY_STRINGS / PROVENANCE_NOT_DECIDED"
    domain_map = pd.DataFrame([
        {"raw_domain": raw, "proposed_canonical_domain": mapped, "record_count": count}
        for (raw, mapped), count in mapping.items()
    ]).sort_values(["proposed_canonical_domain", "raw_domain"])
    frames = {
        "direction_split_domain_full_population": dsd,
        "direction_source_domain_full_population": dsource,
        "duplicate_mechanism_decomposition": mechanisms,
        "duplicate_direction_mirroring_explanation": explanation,
        "duplicate_group_multiplicity": mult,
        "sbs_scope_check": sbs_check,
        "crowd_source_variant_observation": crowd,
        "canonical_domain_mapping_preview": domain_map,
    }
    _save_frames(frames, output_dir)
    configure_plotting()
    direction_matrix = np.array([[within_enko, mirror_rows], [mirror_rows, within_koen]])
    split_matrix = np.array([[within_train, split_rows], [split_rows, within_valid]])
    fig, axes = plt.subplots(1, 2, figsize=(15, 5), constrained_layout=True)
    for axis, matrix, labels, title in [
        (axes[0], direction_matrix, ["EN_TO_KO", "KO_TO_EN"], "Direction decomposition"),
        (axes[1], split_matrix, ["TRAIN", "VALID"], "Split decomposition"),
    ]:
        image = axis.imshow(matrix, cmap="YlOrRd")
        axis.set_xticks(range(2), labels)
        axis.set_yticks(range(2), labels)
        axis.set_title(title)
        for i in range(2):
            for j in range(2):
                axis.text(j, i, f"{matrix[i, j]:,}", ha="center", va="center")
        fig.colorbar(image, ax=axis, label="candidate rows")
    fig.suptitle("AIHub 025 — exact-pair duplicate mechanisms; non-exclusive structural candidates")
    _save_figure(fig, figure_dir / "aihub_025_duplicate_mechanisms.png")
    return frames


def targeted_026(raw_root: Path, output_dir: Path, figure_dir: Path) -> dict[str, pd.DataFrame]:
    counts = Counter()
    mapping = Counter()
    for path in _select_data_files(raw_root, "json"):
        for row in iter_json_data(path):
            source = str(row.get("source") or "<MISSING>")
            domain = str(row.get("domain") or "<MISSING>")
            counts[(source, domain)] += 1
            mapping[(domain, "technology" if domain == "기술과학" else "other")] += 1
    source_domain = pd.DataFrame([
        {"raw_source": source, "raw_domain": domain, "record_count": count}
        for (source, domain), count in counts.items()
    ]).sort_values(["raw_source", "raw_domain"])
    patent = sum(count for (source, _), count in counts.items() if source == "특허정보원")
    technology = sum(count for (_, domain), count in counts.items() if domain == "기술과학")
    joint = counts.get(("특허정보원", "기술과학"), 0)
    check = pd.DataFrame([{
        "patent_source_record_count": patent,
        "technology_domain_record_count": technology,
        "joint_record_count": joint,
        "patent_source_outside_technology_count": patent - joint,
        "technology_domain_outside_patent_source_count": technology - joint,
        "row_level_biconditional_exact": bool(patent == technology == joint),
        "classification": "LOCAL_CATEGORICAL_BICONDITIONAL_ONLY",
    }])
    domain_map = pd.DataFrame([
        {"raw_domain": raw, "proposed_canonical_domain": mapped, "record_count": count}
        for (raw, mapped), count in mapping.items()
    ]).sort_values(["proposed_canonical_domain", "raw_domain"])
    frames = {"source_domain_full_population": source_domain, "patent_technology_row_level_check": check, "canonical_domain_mapping_preview": domain_map}
    _save_frames(frames, output_dir)
    configure_plotting()
    matrix = source_domain.pivot_table(index="raw_source", columns="raw_domain", values="record_count", fill_value=0)
    fig, ax = plt.subplots(figsize=(10, 5), constrained_layout=True)
    image = ax.imshow(matrix.values, aspect="auto", cmap="Blues")
    ax.set_xticks(range(len(matrix.columns)), matrix.columns, rotation=30, ha="right")
    ax.set_yticks(range(len(matrix.index)), matrix.index)
    ax.set_title("AIHub 026 — raw source × raw domain; full population")
    fig.colorbar(image, ax=ax, label="records")
    _save_figure(fig, figure_dir / "aihub_026_source_domain.png")
    return frames


def targeted_legacy(raw_root: Path, output_dir: Path, figure_dir: Path) -> dict[str, pd.DataFrame]:
    news_path = raw_root / "3_문어체_뉴스(2).xlsx"
    culture_path = raw_root / "4_문어체_한국문화.xlsx"
    salt = os.urandom(32)

    def pair_counts(path: Path) -> tuple[Counter[bytes], int]:
        _, rows = iter_xlsx_rows(path)
        counts: Counter[bytes] = Counter()
        total = 0
        for row in rows:
            counts[_pair_digest(str(row.get("원문") or ""), str(row.get("번역문") or ""), salt)] += 1
            total += 1
        return counts, total

    news_counts, news_total = pair_counts(news_path)
    culture_counts, culture_total = pair_counts(culture_path)
    overlap = set(news_counts).intersection(culture_counts)
    news_rows = sum(news_counts[key] for key in overlap)
    culture_rows = sum(culture_counts[key] for key in overlap)
    matchable = sum(min(news_counts[key], culture_counts[key]) for key in overlap)
    summary = pd.DataFrame([{
        "overlap_exact_pair_group_count": len(overlap),
        "one_to_one_matchable_overlap_count": matchable,
        "news2_total_rows": news_total,
        "news2_rows_in_overlap_groups": news_rows,
        "news2_overlap_row_rate": news_rows / max(1, news_total),
        "culture_total_rows": culture_total,
        "culture_rows_in_overlap_groups": culture_rows,
        "culture_overlap_row_rate": culture_rows / max(1, culture_total),
        "classification": "POTENTIAL_CORPUS_COMPOSITION_OVERLAP",
    }])
    multiplicity = pd.DataFrame([
        {"news2_group_multiplicity": pair[0], "culture_group_multiplicity": pair[1], "combined_group_multiplicity": sum(pair), "exact_pair_group_count": count}
        for pair, count in sorted(Counter((news_counts[key], culture_counts[key]) for key in overlap).items())
    ])
    metadata = Counter()
    for corpus, path, fields in [
        ("NEWS2", news_path, ["날짜", "자동분류1", "자동분류2", "자동분류3", "언론사"]),
        ("CULTURE", culture_path, ["키워드"]),
    ]:
        _, rows = iter_xlsx_rows(path)
        for row in rows:
            if _pair_digest(str(row.get("원문") or ""), str(row.get("번역문") or ""), salt) not in overlap:
                continue
            for field in fields:
                value = str(row.get(field) if row.get(field) not in (None, "") else "<MISSING>")
                if len(value) <= 120 and not URL_RE.search(value) and not EMAIL_RE.search(value) and not PHONE_RE.search(value):
                    metadata[(corpus, field, value)] += 1
    metadata_df = pd.DataFrame([
        {"corpus": corpus, "metadata_field": field, "category": value, "overlap_row_count": count}
        for (corpus, field, value), count in metadata.items()
    ]).sort_values(["corpus", "metadata_field", "overlap_row_count"], ascending=[True, True, False])
    concentration_rows = []
    for (corpus, field), part in metadata_df.groupby(["corpus", "metadata_field"]):
        counts = part.overlap_row_count.sort_values(ascending=False)
        denominator = int(counts.sum())
        concentration_rows.append({
            "corpus": corpus,
            "metadata_field": field,
            "overlap_rows_with_field": denominator,
            "unique_category_count": len(counts),
            "top1_share": counts.head(1).sum() / max(1, denominator),
            "top5_share": counts.head(5).sum() / max(1, denominator),
            "top10_share": counts.head(10).sum() / max(1, denominator),
            "classification": "DESCRIPTIVE_CONCENTRATION_ONLY",
        })
    concentration = pd.DataFrame(concentration_rows)
    frames = {
        "news2_culture_overlap_summary": summary,
        "news2_culture_overlap_multiplicity": multiplicity,
        "news2_culture_overlap_metadata_distribution": metadata_df,
        "news2_culture_overlap_concentration": concentration,
    }
    _save_frames(frames, output_dir)
    configure_plotting()
    fig, axes = plt.subplots(1, 2, figsize=(13, 5), constrained_layout=True)
    axes[0].bar(["뉴스(2)", "한국문화"], [summary.iloc[0].news2_overlap_row_rate, summary.iloc[0].culture_overlap_row_rate], color=["#0F766E", "#7C3AED"])
    axes[0].set_title("Rows in overlap groups")
    axes[0].set_ylabel("share of workbook rows")
    mult = multiplicity.groupby("combined_group_multiplicity").exact_pair_group_count.sum()
    axes[1].bar(mult.index.astype(str), mult.values, color="#DC2626")
    axes[1].set_title("Overlap group multiplicity")
    axes[1].set_xlabel("combined group size")
    fig.suptitle("Legacy News2 ↔ Culture — POTENTIAL CORPUS COMPOSITION OVERLAP")
    _save_figure(fig, figure_dir / "legacy_news2_culture_overlap.png")
    return frames


def validate_safe_exports(paths: list[Path]) -> pd.DataFrame:
    rows = []
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(path)
        if path.suffix.lower() in {".csv", ".json", ".md", ".py"}:
            text = path.read_text(encoding="utf-8-sig")
            if URL_RE.search(text) or EMAIL_RE.search(text):
                raise AssertionError(f"URL/email-like export: {path.name}")
            if path.suffix.lower() == ".csv":
                frame = pd.read_csv(path, dtype=str, keep_default_na=False)
                forbidden = RAW_COLUMN_NAMES.intersection(frame.columns)
                if forbidden:
                    raise AssertionError(f"raw columns {sorted(forbidden)}: {path.name}")
                if any("hash" in column.lower() and column.lower() != "sha256" for column in frame.columns):
                    raise AssertionError(f"text-hash-like column: {path.name}")
        rows.append({
            "relative_path": path.as_posix(),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "validation": "SAFE_AGGREGATE_OR_CODE",
        })
    return pd.DataFrame(rows)
