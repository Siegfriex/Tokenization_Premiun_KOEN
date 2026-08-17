#!/usr/bin/env python3
"""Read-only, streaming profiler for the bounded local AIHub raw snapshot."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import mimetypes
import re
import stat
import warnings
import zipfile
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook


HTML_RE = re.compile(r"<[A-Za-z/!][^>\n]{0,200}>")
ENTITY_RE = re.compile(r"&(?:#\d+|#x[0-9A-Fa-f]+|[A-Za-z][A-Za-z0-9]+);")
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
ZERO_WIDTH_RE = re.compile(r"[\u200b\u200c\u200d\u2060\ufeff]")
HANGUL_RE = re.compile(r"[가-힣ㄱ-ㅎㅏ-ㅣ]")
LATIN_RE = re.compile(r"[A-Za-z]")
LENGTH_BINS = ((0, 0), (1, 20), (21, 40), (41, 80), (81, 160), (161, 320), (321, 640), (641, None))


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_stat(path: Path) -> dict[str, int]:
    info = path.stat()
    return {"byte_size": info.st_size, "mtime_epoch_ns": info.st_mtime_ns, "mode": stat.S_IMODE(info.st_mode)}


def json_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return type(value).__name__


def label(value: Any) -> str:
    if value is None:
        return "<NULL>"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def short_hash(parts: list[str]) -> int:
    digest = hashlib.blake2b(digest_size=8)
    for part in parts:
        encoded = part.encode("utf-8", errors="strict")
        digest.update(len(encoded).to_bytes(8, "big"))
        digest.update(encoded)
    return int.from_bytes(digest.digest(), "big")


def iter_json_data(path: Path) -> Iterator[dict[str, Any]]:
    decoder = json.JSONDecoder()
    with path.open("r", encoding="utf-8-sig", errors="strict") as handle:
        prefix = ""
        match = None
        while match is None:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                raise ValueError("top-level data array not found")
            prefix += chunk
            match = re.search(r'"data"\s*:\s*\[', prefix)
        buffer = prefix[match.end():]
        position = 0
        while True:
            while True:
                while position < len(buffer) and buffer[position] in " \t\r\n,":
                    position += 1
                if position < len(buffer):
                    break
                buffer = handle.read(4 * 1024 * 1024)
                position = 0
                if not buffer:
                    raise ValueError("unexpected EOF in data array")
            if buffer[position] == "]":
                return
            while True:
                try:
                    value, end = decoder.raw_decode(buffer, position)
                    break
                except json.JSONDecodeError:
                    remainder = buffer[position:]
                    chunk = handle.read(4 * 1024 * 1024)
                    if not chunk:
                        raise
                    buffer = remainder + chunk
                    position = 0
            if not isinstance(value, dict):
                raise TypeError(f"data array member is {type(value).__name__}, expected object")
            yield value
            position = end
            if position > 8 * 1024 * 1024:
                buffer = buffer[position:]
                position = 0


class ProfileAccumulator:
    def __init__(self, ko_key: str, en_key: str, pair_key: str | None) -> None:
        self.ko_key = ko_key
        self.en_key = en_key
        self.pair_key = pair_key
        self.records = 0
        self.columns: dict[str, dict[str, Any]] = {}
        self.lengths = {ko_key: Counter(), en_key: Counter()}
        self.noise = {ko_key: Counter(), en_key: Counter()}
        self.metadata: dict[str, Counter[str]] = {}
        self.pair_hashes: set[int] = set()
        self.pair_duplicate_candidates = 0
        self.key_hashes: set[int] = set()
        self.key_duplicate_candidates = 0

    def add(self, row: dict[str, Any], metadata_keys: list[str]) -> None:
        self.records += 1
        for key in row:
            if key not in self.columns:
                self.columns[key] = {
                    "absent_count": self.records - 1,
                    "null_count": 0,
                    "empty_string_count": 0,
                    "type_counts": Counter(),
                }
        for key, state in self.columns.items():
            if key not in row:
                state["absent_count"] += 1
                continue
            value = row[key]
            state["type_counts"][json_type(value)] += 1
            if value is None:
                state["null_count"] += 1
            elif isinstance(value, str) and value == "":
                state["empty_string_count"] += 1

        ko = row.get(self.ko_key)
        en = row.get(self.en_key)
        ko_text = "" if ko is None else str(ko)
        en_text = "" if en is None else str(en)
        self._add_text(self.ko_key, ko_text, is_ko=True)
        self._add_text(self.en_key, en_text, is_ko=False)
        pair_digest = short_hash([ko_text, en_text])
        if pair_digest in self.pair_hashes:
            self.pair_duplicate_candidates += 1
        else:
            self.pair_hashes.add(pair_digest)

        if self.pair_key is not None:
            key_value = row.get(self.pair_key)
            if key_value is not None and key_value != "":
                key_digest = short_hash([label(key_value)])
                if key_digest in self.key_hashes:
                    self.key_duplicate_candidates += 1
                else:
                    self.key_hashes.add(key_digest)

        for key in metadata_keys:
            if key in row:
                self.metadata.setdefault(key, Counter())[label(row[key])] += 1

    def _add_text(self, key: str, text: str, is_ko: bool) -> None:
        self.lengths[key][len(text)] += 1
        noise = self.noise[key]
        if text == "":
            noise["empty_rows"] += 1
        if HTML_RE.search(text):
            noise["html_markup_rows"] += 1
        if ENTITY_RE.search(text):
            noise["entity_markup_rows"] += 1
        controls = CONTROL_RE.findall(text)
        if controls:
            noise["control_char_rows"] += 1
            noise["control_char_count"] += len(controls)
        zero_width = ZERO_WIDTH_RE.findall(text)
        if zero_width:
            noise["zero_width_rows"] += 1
            noise["zero_width_count"] += len(zero_width)
        if "\ufffd" in text:
            noise["replacement_char_rows"] += 1
            noise["replacement_char_count"] += text.count("\ufffd")
        if is_ko:
            if LATIN_RE.search(text):
                noise["rows_with_latin"] += 1
            if text and not HANGUL_RE.search(text):
                noise["nonempty_rows_without_hangul"] += 1
        else:
            if HANGUL_RE.search(text):
                noise["rows_with_hangul"] += 1
            if text and not LATIN_RE.search(text):
                noise["nonempty_rows_without_latin"] += 1

    def result(self) -> dict[str, Any]:
        return {
            "record_count": self.records,
            "schema": {
                key: {
                    "absent_count": state["absent_count"],
                    "null_count": state["null_count"],
                    "empty_string_count": state["empty_string_count"],
                    "missing_or_null_count": state["absent_count"] + state["null_count"],
                    "type_counts": dict(sorted(state["type_counts"].items())),
                }
                for key, state in sorted(self.columns.items())
            },
            "raw_length_distribution_chars": {
                key: summarize_lengths(counter) for key, counter in self.lengths.items()
            },
            "markup_control_script_observations": {
                key: dict(sorted(counter.items())) for key, counter in self.noise.items()
            },
            "duplicate_candidates": {
                "pair_hash_method": "BLAKE2b-64 over length-prefixed raw KO and EN strings; collisions are possible, so counts are candidates",
                "duplicate_pair_rows": self.pair_duplicate_candidates,
                "pair_key": self.pair_key,
                "duplicate_nonempty_pair_key_rows": self.key_duplicate_candidates if self.pair_key else None,
            },
            "metadata_value_counts": {
                key: {
                    "distinct_count": len(counter),
                    "top_25": [[value, count] for value, count in counter.most_common(25)],
                }
                for key, counter in sorted(self.metadata.items())
            },
        }


def summarize_lengths(counter: Counter[int]) -> dict[str, Any]:
    total = sum(counter.values())
    if total == 0:
        return {"count": 0}
    ordered = sorted(counter.items())
    def quantile(p: float) -> int:
        target = max(1, math.ceil(total * p))
        cumulative = 0
        for length, count in ordered:
            cumulative += count
            if cumulative >= target:
                return length
        return ordered[-1][0]
    bins: dict[str, int] = {}
    for low, high in LENGTH_BINS:
        name = f"{low}+" if high is None else (str(low) if low == high else f"{low}-{high}")
        bins[name] = sum(count for length, count in ordered if length >= low and (high is None or length <= high))
    weighted = sum(length * count for length, count in ordered)
    return {
        "count": total,
        "min": ordered[0][0],
        "p50": quantile(0.50),
        "p90": quantile(0.90),
        "p95": quantile(0.95),
        "p99": quantile(0.99),
        "max": ordered[-1][0],
        "mean": weighted / total,
        "bins": bins,
    }


def profile_json(path: Path) -> dict[str, Any]:
    accumulator = ProfileAccumulator("ko", "en", "sn")
    metadata = [
        "data_set", "domain", "subdomain", "source_language", "target_language", "file_name",
        "source", "license", "style", "included_unknown_words",
    ]
    for row in iter_json_data(path):
        accumulator.add(row, metadata)
    result = accumulator.result()
    result.update({
        "format": "JSON object with top-level data array",
        "encoding": "UTF-8 (strict full-stream decode)",
        "record_grain": "one translation pair per data[] object",
        "ko_text_candidate": "ko",
        "en_text_candidate": "en",
        "translation_metadata_candidates": ["ko_original", "en_original", "mt", "source_language", "target_language", "word_count_ko", "word_count_en", "word_ratio"],
        "domain_source_metadata_candidates": ["data_set", "domain", "subdomain", "file_name", "source", "license", "style"],
        "pairability": "DIRECT: ko and en coexist in each object; sn is the candidate row key",
    })
    return result


def profile_xlsx(path: Path) -> dict[str, Any]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        raw_header = next(rows, ())
        last_nonempty = max((index for index, value in enumerate(raw_header) if value not in (None, "")), default=-1)
        headers = [str(value) if value not in (None, "") else f"__unnamed_{index + 1}" for index, value in enumerate(raw_header[:last_nonempty + 1])]
        id_key = "SID" if "SID" in headers else ("ID" if "ID" in headers else None)
        accumulator = ProfileAccumulator("원문", "번역문", id_key)
        metadata_keys = [key for key in headers if key not in {"원문", "번역문", id_key}]
        blank_physical_rows = 0
        for values in rows:
            trimmed = values[:len(headers)]
            if not any(value not in (None, "") for value in trimmed):
                blank_physical_rows += 1
                continue
            row = {headers[index]: value for index, value in enumerate(trimmed)}
            accumulator.add(row, metadata_keys)
        sheet_result = accumulator.result()
        sheet_result.update({
            "sheet_name": worksheet.title,
            "worksheet_declared_max_row": worksheet.max_row,
            "worksheet_declared_max_column": worksheet.max_column,
            "blank_physical_rows_after_header": blank_physical_rows,
            "header": headers,
            "record_grain": "one worksheet data row per translation pair",
            "ko_text_candidate": "원문",
            "en_text_candidate": "번역문",
            "pairability": "DIRECT: 원문 and 번역문 coexist in each row",
            "pair_key_assessment": (
                f"{id_key} candidate" if id_key else
                "no explicit unique ID; Set Nr. + 발화자 is a conversation-position candidate, physical row number is deterministic fallback"
            ),
        })
        sheets.append(sheet_result)
    workbook.close()
    return {
        "format": "XLSX (ZIP/XML workbook)",
        "encoding": "not a single text encoding; XML members decoded by openpyxl as Unicode",
        "sheets": sheets,
    }


def zip_inventory(path: Path) -> dict[str, Any]:
    members = []
    with zipfile.ZipFile(path) as archive:
        bad_member = archive.testzip()
        for info in archive.infolist():
            digest = hashlib.sha256()
            with archive.open(info, "r") as handle:
                for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
                    digest.update(chunk)
            members.append({
                "member_path": info.filename,
                "uncompressed_bytes": info.file_size,
                "compressed_bytes": info.compress_size,
                "crc32_hex": f"{info.CRC:08x}",
                "sha256": digest.hexdigest(),
            })
    return {"integrity_test_bad_member": bad_member, "member_count": len(members), "members": members}


def dataset_id(relative_path: str) -> str:
    return relative_path.split("/", 1)[0]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("raw_root", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--branch", required=True)
    parser.add_argument("--head", required=True)
    parser.add_argument("--base-commit", required=True)
    args = parser.parse_args()

    started = now_iso()
    paths = sorted((path for path in args.raw_root.rglob("*") if path.is_file()), key=lambda path: path.relative_to(args.raw_root).as_posix().encode("utf-8"))
    before = {path: file_stat(path) for path in paths}
    hashes = {path: sha256_file(path) for path in paths}
    hash_first_path: dict[str, Path] = {}
    profiles: dict[Path, dict[str, Any]] = {}
    zip_profiles: dict[Path, dict[str, Any]] = {}
    for path in paths:
        suffix = path.suffix.lower()
        digest = hashes[path]
        if suffix in {".json", ".xlsx"} and digest not in hash_first_path:
            hash_first_path[digest] = path
            profiles[path] = profile_json(path) if suffix == ".json" else profile_xlsx(path)
        elif suffix == ".zip":
            zip_profiles[path] = zip_inventory(path)

    after = {path: file_stat(path) for path in paths}
    file_entries = []
    for path in paths:
        rel = path.relative_to(args.raw_root).as_posix()
        unchanged = before[path] == after[path]
        suffix = path.suffix.lower()
        first = hash_first_path.get(hashes[path])
        state = "STABLE_FILE" if unchanged else "DOWNLOAD_IN_PROGRESS"
        entry: dict[str, Any] = {
            "relative_path": rel,
            **after[path],
            "sha256": hashes[path],
            "format": {".json": "JSON", ".xlsx": "XLSX", ".zip": "ZIP"}.get(suffix, suffix.lstrip(".").upper() or "UNKNOWN"),
            "mime_type": mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            "stability_state": state,
            "start_stat": before[path],
            "end_stat": after[path],
            "dataset_id": dataset_id(rel),
        }
        if first is not None and first != path:
            entry["byte_identical_to"] = first.relative_to(args.raw_root).as_posix()
        if unchanged and path in profiles:
            entry["profile"] = profiles[path]
        elif unchanged and first is not None and first in profiles:
            entry["profile_reference"] = first.relative_to(args.raw_root).as_posix()
        if unchanged and path in zip_profiles:
            entry["archive_inventory"] = zip_profiles[path]
        if not unchanged:
            entry["analysis_exclusion_reason"] = "size or mtime changed during audit"
        file_entries.append(entry)

    canonical_lines = [
        f"{entry['relative_path']}\t{entry['byte_size']}\t{entry['mtime_epoch_ns']}\t{entry['sha256']}\n"
        for entry in file_entries
    ]
    raw_manifest_sha = hashlib.sha256("".join(canonical_lines).encode("utf-8")).hexdigest()
    state_counts = Counter(entry["stability_state"] for entry in file_entries)
    datasets: dict[str, dict[str, Any]] = {}
    for entry in file_entries:
        item = datasets.setdefault(entry["dataset_id"], {"physical_file_count": 0, "physical_bytes": 0, "stable_files": 0, "unique_content_sha256_count": set(), "profiled_record_count_unique_content": 0})
        item["physical_file_count"] += 1
        item["physical_bytes"] += entry["byte_size"]
        if entry["stability_state"] == "STABLE_FILE":
            item["stable_files"] += 1
        item["unique_content_sha256_count"].add(entry["sha256"])
        profile = entry.get("profile")
        if profile and entry["format"] == "JSON":
            item["profiled_record_count_unique_content"] += profile["record_count"]
        elif profile and entry["format"] == "XLSX":
            item["profiled_record_count_unique_content"] += sum(sheet["record_count"] for sheet in profile["sheets"])
    for item in datasets.values():
        item["unique_content_sha256_count"] = len(item["unique_content_sha256_count"])

    result = {
        "audit_timestamp": started,
        "audit_completed_timestamp": now_iso(),
        "audit_status": "PROFILED" if paths and not state_counts["DOWNLOAD_IN_PROGRESS"] else ("BLOCKED_NO_LOCAL_RAW_FILES" if not paths else "PROFILED_WITH_DOWNLOAD_IN_PROGRESS_EXCLUSIONS"),
        "raw_root": str(args.raw_root),
        "raw_root_exists": args.raw_root.is_dir(),
        "raw_file_count": len(paths),
        "raw_file_manifest_algorithm": "sha256",
        "raw_file_manifest_canonicalization": "UTF-8 lines sorted by relative_path UTF-8 bytes; relative_path<TAB>byte_size<TAB>mtime_epoch_ns<TAB>sha256<LF>",
        "raw_file_manifest_sha": raw_manifest_sha,
        "branch": args.branch,
        "head": args.head,
        "base_commit": args.base_commit,
        "file_state_counts": {key: state_counts.get(key, 0) for key in ["PARTIAL_DOWNLOAD", "STABLE_FILE", "UNKNOWN", "DOWNLOAD_IN_PROGRESS"]},
        "dataset_summaries": datasets,
        "files": file_entries,
        "method_boundaries": [
            "Raw profiling only; no tokenization, morphology, normalization persistence, regression, inferential statistics, final QC acceptance, or canonical registry creation.",
            "Byte-identical source/label copies are inventoried separately but content-profiled once per SHA-256.",
            "No archive was extracted; ZIP members were streamed read-only for integrity and SHA-256.",
        ],
    }
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
