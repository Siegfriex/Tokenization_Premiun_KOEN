#!/usr/bin/env python3
"""Candidate duplicate and split-overlap profiling over unique raw content."""

from __future__ import annotations

import argparse
import json
import warnings
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from profile_aihub_raw import iter_json_data, short_hash


def json_rows(path: Path) -> Iterator[tuple[str, str, str | None]]:
    for row in iter_json_data(path):
        yield str(row.get("ko") or ""), str(row.get("en") or ""), None if row.get("sn") in (None, "") else str(row["sn"])


def xlsx_rows(path: Path) -> Iterator[tuple[str, str, str | None]]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header_values = next(rows, ())
        headers = [str(value) if value not in (None, "") else f"__unnamed_{index + 1}" for index, value in enumerate(header_values)]
        ko_index = headers.index("원문")
        en_index = headers.index("번역문")
        key_index = headers.index("SID") if "SID" in headers else (headers.index("ID") if "ID" in headers else None)
        for values in rows:
            ko = values[ko_index] if ko_index < len(values) else None
            en = values[en_index] if en_index < len(values) else None
            key = values[key_index] if key_index is not None and key_index < len(values) else None
            if all(value in (None, "") for value in (ko, en, key)):
                continue
            yield str(ko or ""), str(en or ""), None if key in (None, "") else str(key)
    workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("raw_root", type=Path)
    parser.add_argument("profile_manifest", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    manifest = json.loads(args.profile_manifest.read_text(encoding="utf-8"))
    by_dataset: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in manifest["files"]:
        if entry.get("profile") is not None and entry["stability_state"] == "STABLE_FILE":
            by_dataset[entry["dataset_id"]].append(entry)

    results: dict[str, Any] = {}
    for dataset, entries in sorted(by_dataset.items()):
        entries.sort(key=lambda item: ("Validation" in item["relative_path"], item["relative_path"].encode("utf-8")))
        seen_pairs: set[int] = set()
        duplicated_pairs: set[int] = set()
        seen_keys: set[int] = set()
        duplicated_keys: set[int] = set()
        training_pairs: set[int] = set()
        training_keys: set[int] = set()
        total_rows = 0
        duplicate_pair_rows = 0
        duplicate_key_rows = 0
        validation_rows = 0
        validation_pair_overlap_rows = 0
        validation_key_overlap_rows = 0
        file_summaries = []
        for entry in entries:
            relative = entry["relative_path"]
            path = args.raw_root / relative
            is_validation = "/2.Validation/" in f"/{relative}"
            iterator = json_rows(path) if entry["format"] == "JSON" else xlsx_rows(path)
            file_rows = file_dup_pairs = file_dup_keys = file_val_pair_overlap = file_val_key_overlap = 0
            for ko, en, key in iterator:
                total_rows += 1
                file_rows += 1
                pair_digest = short_hash([ko, en])
                if pair_digest in seen_pairs:
                    duplicate_pair_rows += 1
                    file_dup_pairs += 1
                    duplicated_pairs.add(pair_digest)
                else:
                    seen_pairs.add(pair_digest)
                key_digest = short_hash([key]) if key is not None else None
                if key_digest is not None:
                    if key_digest in seen_keys:
                        duplicate_key_rows += 1
                        file_dup_keys += 1
                        duplicated_keys.add(key_digest)
                    else:
                        seen_keys.add(key_digest)
                if is_validation:
                    validation_rows += 1
                    if pair_digest in training_pairs:
                        validation_pair_overlap_rows += 1
                        file_val_pair_overlap += 1
                    if key_digest is not None and key_digest in training_keys:
                        validation_key_overlap_rows += 1
                        file_val_key_overlap += 1
                else:
                    training_pairs.add(pair_digest)
                    if key_digest is not None:
                        training_keys.add(key_digest)
            file_summaries.append({
                "relative_path": relative,
                "record_count": file_rows,
                "is_validation": is_validation,
                "duplicate_pair_rows_against_prior_unique_content": file_dup_pairs,
                "duplicate_nonempty_key_rows_against_prior_unique_content": file_dup_keys,
                "validation_pair_overlap_rows_with_training": file_val_pair_overlap,
                "validation_nonempty_key_overlap_rows_with_training": file_val_key_overlap,
            })
        results[dataset] = {
            "profiled_unique_content_file_count": len(entries),
            "record_count": total_rows,
            "candidate_method": "BLAKE2b-64 over raw, length-prefixed strings; no normalization; hash collisions are possible",
            "duplicate_pair_rows_after_first_occurrence": duplicate_pair_rows,
            "distinct_pair_hashes_with_duplicates": len(duplicated_pairs),
            "duplicate_nonempty_key_rows_after_first_occurrence": duplicate_key_rows,
            "distinct_key_hashes_with_duplicates": len(duplicated_keys),
            "validation_record_count": validation_rows,
            "validation_pair_overlap_rows_with_training": validation_pair_overlap_rows,
            "validation_nonempty_key_overlap_rows_with_training": validation_key_overlap_rows,
            "file_sequence": file_summaries,
        }
    output = {
        "source_profile_manifest": args.profile_manifest.name,
        "source_raw_file_manifest_sha": manifest["raw_file_manifest_sha"],
        "raw_root": str(args.raw_root),
        "datasets": results,
        "scope_note": "Candidate duplicate profiling only; this is not QC acceptance or inferential analysis.",
    }
    args.output.write_text(json.dumps(output, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
