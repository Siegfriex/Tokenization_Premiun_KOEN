"""G1 D-01 pair registry를 exact allowlist에서 streaming 구축·검증하는 engineering 계층이다."""

from __future__ import annotations  # Python 3.12 지연 annotation 평가를 사용한다.

import csv  # human-readable reconciliation table을 UTF-8 CSV로 저장한다.
import datetime as dt  # workbook 날짜 값을 손실 없는 ISO 문자열로 보존한다.
import hashlib  # provenance identity와 exact raw content identity의 SHA-256을 계산한다.
import json  # raw locator와 metadata를 canonical JSON 문자열로 보존한다.
import os  # 완성된 artifact를 atomic replace한다.
import re  # DuckDB memory override를 제한된 size 문법으로 검증한다.
import struct  # identity 문자열의 UTF-8 byte 길이를 unsigned 64-bit big-endian으로 직렬화한다.
import subprocess  # artifact에 연결할 Git commit을 read-only 조회한다.
from collections.abc import Iterator, Mapping, Sequence  # streaming record 및 contract container 타입을 표현한다.
from pathlib import Path  # project/raw/output 경로를 명시적으로 처리한다.
from typing import Any  # heterogeneous raw JSON/XLSX 값을 타입으로 표현한다.

import duckdb  # 565만 행 duplicate group window와 reconciliation을 out-of-core SQL로 계산한다.
import ijson  # 대용량 AIHub JSON data 배열을 전체 메모리 적재 없이 읽는다.
import openpyxl  # Legacy XLSX를 read-only streaming mode로 읽는다.
import pyarrow as pa  # 명시적 schema의 record batch를 구성한다.
import pyarrow.parquet as pq  # streaming staging 및 canonical Parquet artifact를 저장한다.
import yaml  # Claude가 freeze한 research config를 읽는다.

from tokenization_premium.hashing import canonical_json_bytes, sha256_file  # project 공통 canonical JSON과 파일 SHA-256 규칙을 재사용한다.
from tokenization_premium.io import write_json  # manifest를 UTF-8 atomic JSON으로 저장한다.
from tokenization_premium.schemas import pair_registry_schema, source_registry_schema, staging_registry_schema  # D-01과 source schema를 단일 정의에서 소비한다.

PAIR_VERSION = "v001"  # SSOT §38의 zero-padded registry version을 고정한다.
PAIR_REGISTRY_SCHEMA_VERSION = "PAIR_REGISTRY_v001"  # manifest에 기록할 pair schema version을 고정한다.
SOURCE_REGISTRY_SCHEMA_VERSION = "SOURCE_REGISTRY_v001"  # manifest에 기록할 source schema version을 고정한다.
BATCH_SIZE = 25_000  # streaming writer의 bounded-memory row batch 크기를 고정한다.
DEFAULT_DUCKDB_MEMORY_LIMIT = "8GB"  # INC-001 이후 15GiB WSL 환경에 맞춘 보수적 기본 상한을 고정한다.
DUCKDB_MEMORY_LIMIT_ENV = "TOKENIZATION_PREMIUM_DUCKDB_MEMORY_LIMIT"  # 명시적 operator override 환경변수 이름을 고정한다.
_DUCKDB_MEMORY_LIMIT_PATTERN = re.compile(r"(?P<amount>[1-9][0-9]*)\s*(?P<unit>KB|MB|GB|TB)", re.IGNORECASE)  # SQL에 안전한 양의 정수 memory-size 문법만 허용한다.
SOURCE_ID_LABELS = {"025": "025-family", "026": "026-family", "LEGACY": "Legacy-family"}  # 공식 dataSetSn을 조작하지 않는 local family label을 고정한다.
LEGACY_ROLE_DOMAIN = {  # workbook role에서 Director 승인 raw top-level label을 결정한다.
    "LEGACY_1_구어체(1)": "구어체",  # 첫 구어체 workbook을 general mapping의 raw label에 연결한다.
    "LEGACY_1_구어체(2)": "구어체",  # 둘째 구어체 workbook을 같은 raw label에 연결한다.
    "LEGACY_2_대화체": "대화체",  # 대화체 workbook을 dialogue mapping의 raw label에 연결한다.
    "LEGACY_3_문어체_뉴스(1)_200226": "뉴스",  # 첫 뉴스 workbook을 news mapping의 raw label에 연결한다.
    "LEGACY_3_문어체_뉴스(2)": "뉴스",  # 둘째 뉴스 workbook을 news mapping의 raw label에 연결한다.
    "LEGACY_3_문어체_뉴스(3)": "뉴스",  # 셋째 뉴스 workbook을 news mapping의 raw label에 연결한다.
    "LEGACY_3_문어체_뉴스(4)": "뉴스",  # 넷째 뉴스 workbook을 news mapping의 raw label에 연결한다.
    "LEGACY_4_문어체_한국문화": "한국문화",  # 한국문화 workbook을 other mapping의 raw label에 연결한다.
    "LEGACY_5_문어체_조례": "조례",  # 조례 workbook을 legal mapping의 raw label에 연결한다.
    "LEGACY_6_문어체_지자체웹사이트": "지자체웹사이트",  # 지자체 workbook을 administration mapping의 raw label에 연결한다.
}


def _jsonable(value: Any) -> Any:
    """
    /**
     * @purpose XLSX/JSON raw value를 의미를 숨기지 않는 JSON 직렬화 가능 값으로 바꾼다.
     * @spec_ref SSOT §11, §30.2; G1_PAIR_REGISTRY_PRECONTRACT_v1 §6
     * @param value source-native scalar/container
     * @return canonical JSON에 넣을 수 있는 scalar/container
     * @raises TypeError 지원하지 않는 source-native 객체가 들어온 경우
     * @validation datetime ISO 보존과 mapping key 정렬을 단위 테스트한다.
     * @artifact raw_metadata_json, raw_locator
     */
    """
    if isinstance(value, Mapping):  # raw mapping이면 key를 문자열화하고 내부 값도 재귀 변환한다.
        return {str(key): _jsonable(item) for key, item in value.items()}  # source field 이름과 값을 빠짐없이 보존한다.
    if isinstance(value, (list, tuple)):  # raw sequence이면 원래 순서를 유지한다.
        return [_jsonable(item) for item in value]  # 각 원소를 JSON 가능 값으로 변환한다.
    if isinstance(value, (dt.datetime, dt.date, dt.time)):  # workbook 날짜/시간이면 locale 비의존 표현이 필요하다.
        return value.isoformat()  # ISO-8601 문자열로 손실 없이 보존한다.
    if value is None or isinstance(value, (str, int, float, bool)):  # JSON native scalar이면 변형할 이유가 없다.
        return value  # 원래 scalar를 그대로 반환한다.
    raise TypeError(f"지원하지 않는 raw metadata 타입: {type(value)!r}")  # silent string coercion을 금지하고 schema drift를 즉시 중단한다.


def canonical_json_text(payload: Any) -> str:
    """
    /**
     * @purpose raw locator/metadata를 key-sorted compact UTF-8 JSON 문자열로 직렬화한다.
     * @spec_ref SSOT §30.2, §38
     * @param payload JSON-compatible 또는 _jsonable로 변환 가능한 객체
     * @return trailing newline이 없는 canonical JSON 문자열
     * @raises TypeError 지원하지 않는 값 또는 JSON 직렬화 실패 시
     * @validation 동일 mapping의 key 순서가 달라도 같은 문자열인지 검사한다.
     * @artifact raw_locator, raw_metadata_json, source input_files_json
     */
    """
    return json.dumps(_jsonable(payload), ensure_ascii=False, sort_keys=True, separators=(",", ":"))  # platform·locale 비의존 canonical JSON을 만든다.


def length_prefixed_sha256(values: Sequence[str]) -> str:
    """
    /**
     * @purpose 문자열 sequence를 u64be 길이-prefix UTF-8 직렬화한 SHA-256으로 식별한다.
     * @spec_ref G1 pair identity contract, D-01 pair_id engineering rule
     * @param values 순서가 의미를 갖는 문자열 sequence
     * @return 64자리 소문자 SHA-256 hex digest
     * @raises TypeError 값이 문자열이 아닌 경우
     * @validation prefix ambiguity 사례와 hashlib reference를 단위 테스트한다.
     * @artifact pair_id, duplicate_group_id
     */
    """
    digest = hashlib.sha256()  # 각 field의 길이와 bytes를 순서대로 누적할 digest를 만든다.
    for value in values:  # identity 구성 field를 계약 순서대로 처리한다.
        if not isinstance(value, str):  # 문자열 외 암묵 변환은 provenance identity drift를 만들 수 있다.
            raise TypeError(f"identity 구성값은 str이어야 한다: {type(value)!r}")  # 잘못된 타입을 fail-fast한다.
        encoded = value.encode("utf-8")  # Unicode text를 변경하지 않고 UTF-8 bytes로 변환한다.
        digest.update(struct.pack(">Q", len(encoded)))  # field boundary를 8-byte big-endian 길이로 고정한다.
        digest.update(encoded)  # 길이 뒤에 원본 UTF-8 bytes를 순서대로 추가한다.
    return digest.hexdigest()  # 비교 가능한 64자리 digest를 반환한다.


def duplicate_group_id(ko_text_raw: str, en_text_raw: str) -> str:
    """
    /**
     * @purpose Recon oracle과 동일한 raw KO+EN exact content identity를 계산한다.
     * @spec_ref G1_INGEST_EXPECTATIONS_v001 duplicate_baseline_oracle.identity_method
     * @param ko_text_raw 정규화하지 않은 한국어 문자열
     * @param en_text_raw 정규화하지 않은 영어 문자열
     * @return u64be(len(KO))||KO||u64be(len(EN))||EN의 SHA-256
     * @raises TypeError 입력이 문자열이 아닌 경우
     * @validation independent reference vector와 oracle counts로 검사한다.
     * @artifact duplicate_group_id
     */
    """
    return length_prefixed_sha256((ko_text_raw, en_text_raw))  # 다른 metadata를 섞지 않고 raw content 두 필드만 hash한다.


def provenance_pair_id(source_id: str, source_record_id: str) -> str:
    """
    /**
     * @purpose source family record provenance로 project-wide pair primary key를 만든다.
     * @spec_ref SSOT §12.1; PAIR_IDENTITY_AND_DUPLICATE_CONTRACT_v1 §1-3
     * @param source_id family와 raw snapshot을 포함한 source identity
     * @param source_record_id source 범위 내 deterministic record identity
     * @return pair_ prefix와 SHA-256 digest
     * @raises TypeError 입력이 문자열이 아닌 경우
     * @validation 전체 registry pair_id uniqueness=100%로 검사한다.
     * @artifact pair_id
     */
    """
    return "pair_" + length_prefixed_sha256((source_id, source_record_id))  # content와 분리된 provenance identity임을 prefix로 명시한다.


def source_id_for(logical_corpus: str, raw_manifest_sha256: str) -> str:
    """
    /**
     * @purpose 확인된 local family와 ingest snapshot을 결합하고 official dataSetSn은 조작하지 않는다.
     * @spec_ref G1_PAIR_REGISTRY_PRECONTRACT_v1 §1, §6a
     * @param logical_corpus 025, 026, LEGACY 중 하나
     * @param raw_manifest_sha256 verified raw manifest SHA-256
     * @return family@raw-manifest:<64hex> source_id
     * @raises KeyError 알 수 없는 family인 경우
     * @validation 3개 family만 생성되고 snapshot hash가 exact 포함되는지 검사한다.
     * @artifact source_id, SOURCE_REGISTRY_v001
     */
    """
    label = SOURCE_ID_LABELS[logical_corpus]  # 공식 ID가 아닌 Director 승인 family label을 선택한다.
    return f"{label}@raw-manifest:{raw_manifest_sha256}"  # snapshot 변경 시 identity가 바뀌도록 전체 manifest SHA를 포함한다.


def load_contracts(research_config_path: Path, ingest_contract_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    """
    /**
     * @purpose Claude research contract와 data-recon ingest oracle을 변경 없이 읽고 필수 G1 상태를 검증한다.
     * @spec_ref D-01 G1 contract dependency directive
     * @param research_config_path research/g1-claude configs/research_v1.yaml
     * @param ingest_contract_path data/g1-recon G1_INGEST_EXPECTATIONS_v001.json
     * @return research config와 ingest contract mapping tuple
     * @raises AssertionError 필수 key/value가 준비되지 않은 경우
     * @validation exact expected counts, Tier, direction, pair version 계약을 assert한다.
     * @artifact 실행 manifest contract hashes
     */
    """
    research = yaml.safe_load(research_config_path.read_text(encoding="utf-8"))  # Claude-owned YAML을 read-only로 파싱한다.
    ingest = json.loads(ingest_contract_path.read_text(encoding="utf-8"))  # data-recon JSON을 read-only로 파싱한다.
    assert ingest["expected_logical_record_counts"] == {"025": 2_700_345, "026": 1_350_162, "LEGACY": 1_602_418, "total": 5_652_925}  # Director가 준 exact logical row oracle을 확인한다.
    assert research["source_hierarchy"]["corpus_tier_assignment"] == {"025": "A", "026": "A", "legacy": None}  # D-RD-05 Tier bundle을 확인한다.
    assert research["translation_direction_defaults"]["legacy"] == "UNKNOWN"  # D-RD-06 Legacy 방향을 확인한다.
    assert research["d01_field_contract"]["pair_quality_status"]["phase1_value"] == "review"  # Phase 1 accepted 오판을 방지한다.
    assert research["d01_field_contract"]["pair_version"]["syntax"].startswith("v001")  # registry version을 config에서 확인한다.
    return research, ingest  # 검증을 통과한 두 계약 mapping을 반환한다.


def verify_ingest_allowlist(raw_root: Path, ingest_contract: Mapping[str, Any]) -> list[dict[str, Any]]:
    """
    /**
     * @purpose exact canonical allowlist 파일의 존재·SHA-256·논리 row sum을 ingest 전에 검증한다.
     * @spec_ref G1_INGEST_EXPECTATIONS_v001 double_ingest_prevention_contract
     * @param raw_root local immutable AIHub raw root
     * @param ingest_contract data-recon expectations mapping
     * @return canonical ingest entry 목록
     * @raises AssertionError 누락, hash mismatch, alias 포함, row sum mismatch 시
     * @validation 16개 exact path와 3개 corpus expected logical rows를 검사한다.
     * @artifact manifest.input_files
     */
    """
    entries = list(ingest_contract["double_ingest_prevention_contract"]["ingest_allowlist"])  # recursive discovery 대신 승인된 exact 목록만 가져온다.
    assert len(entries) == 16  # 현재 freeze된 6 JSON+10 XLSX canonical physical file 수를 확인한다.
    observed_sums = {"025": 0, "026": 0, "LEGACY": 0}  # family별 contract row sum을 누적한다.
    seen_paths: set[str] = set()  # 동일 physical path가 allowlist에 중복되지 않았는지 추적한다.
    for entry in entries:  # 모든 canonical entry를 한 번씩만 검증한다.
        relative_path = str(entry["relative_path"])  # contract의 machine-independent relative path를 읽는다.
        assert relative_path not in seen_paths  # 동일 경로의 accidental double ingest를 fail-fast한다.
        seen_paths.add(relative_path)  # 확인한 경로를 중복 검사용 set에 기록한다.
        path = raw_root / relative_path  # contract raw root 아래에서만 physical file을 해석한다.
        assert path.is_file(), f"raw allowlist file 누락: {path}"  # 누락 파일이 있으면 ingest를 시작하지 않는다.
        observed_sha256 = sha256_file(path)  # 실제 bytes의 SHA-256을 streaming 계산한다.
        assert observed_sha256 == entry["sha256"], f"raw SHA mismatch: {relative_path}"  # data-recon hash와 다르면 중단한다.
        observed_sums[str(entry["logical_corpus"])] += int(entry["record_count"])  # contract logical rows를 family별 누적한다.
    expected = dict(ingest_contract["expected_logical_record_counts"])  # independent expected family counts를 읽는다.
    assert observed_sums == {key: int(expected[key]) for key in ("025", "026", "LEGACY")}  # allowlist row 합이 exact oracle과 같은지 확인한다.
    return entries  # 검증된 exact canonical entry만 ingest에 전달한다.


def _raw_metadata_without_text(record: Mapping[str, Any], text_keys: set[str]) -> str:
    """
    /**
     * @purpose raw text는 D-01 columns에 두고 나머지 source row metadata 전체를 canonical JSON으로 보존한다.
     * @spec_ref SSOT §11; G1_PAIR_REGISTRY_PRECONTRACT_v1 §6
     * @param record source row mapping
     * @param text_keys 별도 D-01 raw text columns에 이미 보존된 key 집합
     * @return text 외 raw fields의 canonical JSON 문자열
     * @raises TypeError 지원하지 않는 source value 타입인 경우
     * @validation 원래 key 집합이 text keys와 metadata keys의 합인지 검사한다.
     * @artifact raw_metadata_json
     */
    """
    metadata = {str(key): value for key, value in record.items() if str(key) not in text_keys and key is not None}  # text 중복을 피하되 나머지 raw field를 모두 유지한다.
    return canonical_json_text(metadata)  # mapping을 deterministic compact UTF-8 JSON으로 변환한다.


def _direction_from_json(record: Mapping[str, Any], expected_direction: str) -> tuple[str, bool]:
    """
    /**
     * @purpose raw original-field presence와 language metadata로 record-level 방향을 계약대로 판정한다.
     * @spec_ref D-RD-06; DIRECTION_AND_DOMAIN_MAPPING_PRECONTRACT_v1 Task 5
     * @param record 025/026 raw JSON data object
     * @param expected_direction allowlist file의 provenance direction
     * @return translation_direction_raw와 review flag
     * @raises AssertionError unknown expected direction인 경우
     * @validation observed raw signals 일치, 약한 fallback, contradiction 사례를 검사한다.
     * @artifact translation_direction_raw, translation_direction_review_flag
     */
    """
    assert expected_direction in {"KO_TO_EN", "EN_TO_KO"}  # 현재 JSON track에서 허용된 두 방향만 받는다.
    source_language = record.get("source_language")  # raw source language signal을 변형 없이 읽는다.
    target_language = record.get("target_language")  # raw target language signal을 변형 없이 읽는다.
    has_ko_original = isinstance(record.get("ko_original"), str) and record.get("ko_original") != ""  # KO original field가 명시적으로 존재하는지 확인한다.
    has_en_original = isinstance(record.get("en_original"), str) and record.get("en_original") != ""  # EN original field가 명시적으로 존재하는지 확인한다.
    language_direction = "KO_TO_EN" if (source_language, target_language) == ("ko", "en") else "EN_TO_KO" if (source_language, target_language) == ("en", "ko") else None  # raw language pair가 지시하는 방향을 계산한다.
    original_direction = "KO_TO_EN" if has_ko_original and not has_en_original else "EN_TO_KO" if has_en_original and not has_ko_original else None  # original field presence가 지시하는 방향을 계산한다.
    if original_direction is not None and language_direction == original_direction == expected_direction:  # 강한 두 raw signal과 file provenance가 모두 일치한다.
        return expected_direction, False  # review가 필요 없는 record-level 방향을 반환한다.
    if original_direction is None and language_direction == expected_direction:  # original field가 없지만 language metadata와 file convention이 일치한다.
        return expected_direction, True  # 계약상 weaker fallback이므로 review flag를 켠다.
    return "UNKNOWN", True  # 모순 또는 불완전 signal을 조용히 선택하지 않고 UNKNOWN review로 보낸다.


def _legacy_auxiliary(record: Mapping[str, Any], domain_raw: str) -> tuple[str | None, str | None, str | None, str | None]:
    """
    /**
     * @purpose Legacy workbook의 raw ID/SID, subdomain, fine-grained source label을 보존한다.
     * @spec_ref G1_PAIR_REGISTRY_PRECONTRACT_v1 §5-6a, §14
     * @param record worksheet header-to-value mapping
     * @param domain_raw approved workbook-level raw domain label
     * @return source_native_id, source_native_sid, subdomain_raw, source_provenance_raw tuple
     * @raises 없음
     * @validation workbook별 sample/header test로 검사한다.
     * @artifact Legacy auxiliary columns
     */
    """
    native_id = record.get("ID")  # 뉴스/문화/조례/지자체 native ID를 읽는다.
    native_sid = record.get("SID")  # 구어체 native SID를 읽는다.
    subdomain = record.get("소분류") if domain_raw == "대화체" else record.get("자동분류1") if domain_raw == "뉴스" else record.get("키워드") if domain_raw == "한국문화" else None  # 명시된 source subdomain field만 원문 그대로 선택한다.
    provenance = record.get("언론사") if domain_raw == "뉴스" else record.get("지자체") if domain_raw in {"조례", "지자체웹사이트"} else None  # publisher/municipality raw source label을 보존한다.
    return _string_or_none(native_id), _string_or_none(native_sid), _string_or_none(subdomain), _string_or_none(provenance)  # nullable string columns에 맞게 값만 형식화한다.


def _string_or_none(value: Any) -> str | None:
    """
    /**
     * @purpose source-native scalar ID/label을 안정적인 문자열로 보존하고 missing은 null로 유지한다.
     * @spec_ref G1 source_record_id_rules.LEGACY
     * @param value workbook scalar
     * @return stable string 또는 None
     * @raises TypeError 비-scalar 값인 경우
     * @validation integer-valued float가 불필요한 .0 없이 보존되는지 검사한다.
     * @artifact source_native_id, source_native_sid, raw labels
     */
    """
    if value is None:  # workbook blank cell은 missing metadata다.
        return None  # 빈 문자열을 조작하지 않고 null로 유지한다.
    if isinstance(value, float) and value.is_integer():  # Excel numeric ID가 1.0으로 읽힐 수 있다.
        return str(int(value))  # semantic integer ID를 1로 안정화한다.
    if isinstance(value, (str, int, float, bool)):  # source-native scalar는 문자열 column에 보존 가능하다.
        return str(value)  # locale 비의존 기본 표현으로 변환한다.
    raise TypeError(f"문자열 보존 불가 scalar: {type(value)!r}")  # container/object의 silent coercion을 금지한다.


def _base_staging_row(
    *,
    source_id: str,
    source_tier: str | None,
    domain: str,
    direction_raw: str,
    direction_review: bool,
    ko_text: str,
    en_text: str,
    source_license_note: str,
    source_record_id: str,
    raw_locator: str,
    domain_raw: str | None,
    subdomain_raw: str | None,
    source_provenance_raw: str | None,
    is_validation: bool,
    mt_field_present: bool,
    logical_corpus: str,
    entry: Mapping[str, Any],
    raw_sheet_name: str | None,
    raw_physical_row_number: int | None,
    source_native_id: str | None,
    source_native_sid: str | None,
    raw_metadata_json: str,
) -> dict[str, Any]:
    """
    /**
     * @purpose JSON/XLSX 공통 Phase-1 staging row를 D-01 계약값으로 조립한다.
     * @spec_ref SSOT §12.1; G1_PAIR_REGISTRY_PRECONTRACT_v1 §14-16
     * @param source_id 외 keyword-only fields 한 raw record의 provenance와 raw content
     * @return staging_registry_schema와 일치하는 row mapping
     * @raises AssertionError raw KO/EN이 구조적으로 유효하지 않거나 schema key가 다른 경우
     * @validation 1 raw structurally valid record = 1 staging row, exact schema keys를 검사한다.
     * @artifact PAIR_REGISTRY_v001 staging rows
     */
    """
    assert isinstance(ko_text, str) and ko_text != ""  # Phase 1 structural ingest에 필요한 KO nonempty string을 확인한다.
    assert isinstance(en_text, str) and en_text != ""  # Phase 1 structural ingest에 필요한 EN nonempty string을 확인한다.
    row = {  # D-01과 승인 auxiliary fields의 Phase-1 값을 한 행으로 구성한다.
        "pair_id": provenance_pair_id(source_id, source_record_id),  # content가 아닌 provenance identity를 계산한다.
        "source_id": source_id,  # family/snapshot source identity를 연결한다.
        "source_tier": source_tier,  # Director 승인 Tier 또는 null을 저장한다.
        "domain": domain,  # top-level mapping으로 정한 canonical domain을 저장한다.
        "sentence_type": "other",  # 현재 raw schema에 explicit sentence type이 없다는 계약값을 적용한다.
        "translation_direction_raw": direction_raw,  # group resolution 전 raw record 방향을 보존한다.
        "translation_direction_review_flag": bool(direction_review),  # weaker/contradictory direction evidence를 표시한다.
        "ko_text_raw": ko_text,  # raw KO text를 어떠한 normalization도 없이 저장한다.
        "en_text_raw": en_text,  # raw EN text를 어떠한 normalization도 없이 저장한다.
        "ko_text_nfc": None,  # Phase 2 전 normalization field를 null로 둔다.
        "en_text_nfc": None,  # Phase 2 전 normalization field를 null로 둔다.
        "ko_text_analysis": None,  # Phase 2 전 analysis text를 null로 둔다.
        "en_text_analysis": None,  # Phase 2 전 analysis text를 null로 둔다.
        "pair_quality_status": "review",  # semantic QC 전에는 accepted를 부여하지 않는다.
        "pair_quality_score": None,  # semantic alignment score 미측정을 null로 표현한다.
        "pair_version": PAIR_VERSION,  # v001 registry version을 연결한다.
        "source_license_note": source_license_note,  # raw license evidence의 한계를 명시한다.
        "source_record_id": source_record_id,  # source 범위 내 deterministic record key를 저장한다.
        "raw_locator": raw_locator,  # physical provenance를 canonical JSON으로 저장한다.
        "duplicate_group_id": duplicate_group_id(ko_text, en_text),  # exact raw KO+EN content digest를 계산한다.
        "domain_raw": domain_raw,  # raw top-level domain label을 변형 없이 저장한다.
        "subdomain_raw": subdomain_raw,  # raw subdomain을 변형 없이 저장한다.
        "source_provenance_raw": source_provenance_raw,  # raw source/publisher label을 변형 없이 저장한다.
        "is_validation_upstream": bool(is_validation),  # upstream split provenance만 저장한다.
        "mt_field_present": bool(mt_field_present),  # raw mt key presence를 저장한다.
        "sentence_type_raw": None,  # current schemas에 explicit sentence type이 없음을 null로 보존한다.
        "sentence_type_provenance_status": "UNAVAILABLE_IN_RAW_SOURCE",  # sentence type 비가용성을 명시한다.
        "normalization_status": "NOT_GENERATED_PHASE1",  # normalization 미실행을 명시한다.
        "qc_stage_status": "PENDING_PHASE2",  # QC 미실행을 명시한다.
        "logical_corpus": logical_corpus,  # family-level reconciliation label을 저장한다.
        "canonical_ingest_role": str(entry["canonical_ingest_role"]),  # exact allowlist role을 연결한다.
        "raw_file_relative_path": str(entry["relative_path"]),  # immutable raw relative path를 연결한다.
        "raw_file_sha256": str(entry["sha256"]),  # 사전 검증된 physical file SHA를 연결한다.
        "raw_sheet_name": raw_sheet_name,  # XLSX sheet 또는 JSON null을 저장한다.
        "raw_physical_row_number": raw_physical_row_number,  # XLSX physical row 또는 JSON null을 저장한다.
        "source_native_id": source_native_id,  # Legacy native ID를 보존한다.
        "source_native_sid": source_native_sid,  # Legacy native SID를 보존한다.
        "raw_metadata_json": raw_metadata_json,  # text 외 raw metadata 전체를 canonical JSON으로 저장한다.
    }
    assert set(row) == set(staging_registry_schema().names)  # 누락/추가 key로 schema가 암묵 drift하지 않게 한다.
    return row  # schema 검증을 통과한 staging row를 반환한다.


def iter_json_rows(
    raw_root: Path,
    entry: Mapping[str, Any],
    research: Mapping[str, Any],
    raw_manifest_sha256: str,
) -> Iterator[dict[str, Any]]:
    """
    /**
     * @purpose 025/026 data[]를 streaming하며 one raw object→one staging row로 변환한다.
     * @spec_ref D-01 ingest contract; D-RD-05~07
     * @param raw_root immutable raw root
     * @param entry exact canonical JSON allowlist entry
     * @param research Claude-frozen research config
     * @param raw_manifest_sha256 source snapshot identity hash
     * @return staging row iterator
     * @raises AssertionError sn/text/domain/schema가 계약을 위반한 경우
     * @validation file별 yielded count와 entry.record_count exact equality를 검사한다.
     * @artifact PAIR_REGISTRY_v001 staging rows
     */
    """
    logical_corpus = str(entry["logical_corpus"])  # 025 또는 026 family label을 읽는다.
    assert logical_corpus in {"025", "026"}  # JSON reader가 Legacy에 잘못 사용되지 않게 한다.
    source_id = source_id_for(logical_corpus, raw_manifest_sha256)  # family와 raw snapshot source identity를 만든다.
    source_tier = research["source_hierarchy"]["corpus_tier_assignment"][logical_corpus]  # Director 승인 Tier A를 읽는다.
    expected_direction = str(entry["canonical_ingest_role"]).split("_", 1)[1].rsplit("_", 1)[0]  # role에서 EN_TO_KO 또는 KO_TO_EN 부분을 분리한다.
    is_validation = str(entry["canonical_ingest_role"]).endswith("_VALIDATION")  # upstream split provenance를 role에서 exact 결정한다.
    path = raw_root / str(entry["relative_path"])  # exact allowlist relative path만 연다.
    count = 0  # file별 실제 yielded rows를 센다.
    with path.open("rb") as stream:  # JSON bytes를 newline/encoding 변환 없이 streaming으로 연다.
        for record in ijson.items(stream, "data.item", use_float=True):  # top-level data array object를 하나씩 읽는다.
            assert isinstance(record, Mapping)  # 각 data item이 object인지 구조적으로 확인한다.
            sn = record.get("sn")  # 025/026 canonical source record key를 읽는다.
            assert isinstance(sn, str) and sn != ""  # data-recon이 보장한 nonempty sn을 독립 확인한다.
            ko_text = record.get("ko")  # paired KO raw field를 읽는다.
            en_text = record.get("en")  # paired EN raw field를 읽는다.
            assert isinstance(ko_text, str) and ko_text != ""  # one structurally valid record의 KO 조건을 확인한다.
            assert isinstance(en_text, str) and en_text != ""  # one structurally valid record의 EN 조건을 확인한다.
            domain_raw = record.get("domain")  # raw top-level domain label을 읽는다.
            assert isinstance(domain_raw, str) and domain_raw in research["domain_mapping_top_level"][logical_corpus]  # 승인 mapping 밖 label을 fail-fast한다.
            domain = research["domain_mapping_top_level"][logical_corpus][domain_raw]  # Director 승인 canonical domain을 적용한다.
            direction_raw, direction_review = _direction_from_json(record, expected_direction)  # raw direction signals를 계약대로 판정한다.
            raw_locator = canonical_json_text({"relative_path": str(entry["relative_path"]), "source_record_id": sn})  # JSON record를 file+sn으로 추적 가능하게 만든다.
            raw_license = record.get("license")  # raw self-asserted license label을 읽는다.
            license_note = f"SELF_ASSERTED_RAW_LICENSE={raw_license!s}; NOT_VERIFIED_AS_REDISTRIBUTION_PERMISSION"  # open을 재배포 승인으로 승격하지 않는다.
            yield _base_staging_row(  # 변형 없는 text와 명시적 Phase-1 상태를 공통 row로 만든다.
                source_id=source_id,
                source_tier=source_tier,
                domain=str(domain),
                direction_raw=direction_raw,
                direction_review=direction_review,
                ko_text=ko_text,
                en_text=en_text,
                source_license_note=license_note,
                source_record_id=sn,
                raw_locator=raw_locator,
                domain_raw=domain_raw,
                subdomain_raw=_string_or_none(record.get("subdomain")),
                source_provenance_raw=_string_or_none(record.get("source")),
                is_validation=is_validation,
                mt_field_present="mt" in record,
                logical_corpus=logical_corpus,
                entry=entry,
                raw_sheet_name=None,
                raw_physical_row_number=None,
                source_native_id=None,
                source_native_sid=None,
                raw_metadata_json=_raw_metadata_without_text(record, {"ko", "en"}),
            )
            count += 1  # 성공적으로 만든 raw record 수를 증가시킨다.
    assert count == int(entry["record_count"]), f"JSON record count mismatch: {entry['relative_path']} expected={entry['record_count']} observed={count}"  # file 종료 시 exact count를 확인한다.


def iter_xlsx_rows(
    raw_root: Path,
    entry: Mapping[str, Any],
    research: Mapping[str, Any],
    raw_manifest_sha256: str,
) -> Iterator[dict[str, Any]]:
    """
    /**
     * @purpose Legacy workbook를 read-only streaming하며 physical locator 기반 staging rows로 변환한다.
     * @spec_ref D-01 ingest contract; source_record_id_rules.LEGACY; D-RD-06~07
     * @param raw_root immutable raw root
     * @param entry exact canonical XLSX allowlist entry
     * @param research Claude-frozen research config
     * @param raw_manifest_sha256 source snapshot identity hash
     * @return staging row iterator
     * @raises AssertionError sheet/header/text/count/domain 계약 위반 시
     * @validation workbook별 yielded count와 entry.record_count exact equality를 검사한다.
     * @artifact PAIR_REGISTRY_v001 staging rows
     */
    """
    logical_corpus = str(entry["logical_corpus"])  # Legacy family label을 읽는다.
    assert logical_corpus == "LEGACY"  # XLSX reader가 JSON family에 잘못 사용되지 않게 한다.
    role = str(entry["canonical_ingest_role"])  # workbook-specific canonical ingest role을 읽는다.
    domain_raw = LEGACY_ROLE_DOMAIN[role]  # 승인된 workbook-level raw top-level domain을 선택한다.
    domain = research["domain_mapping_top_level"]["legacy"][domain_raw]  # Director 승인 canonical domain을 적용한다.
    source_id = source_id_for(logical_corpus, raw_manifest_sha256)  # Legacy family와 raw snapshot identity를 만든다.
    source_tier = research["source_hierarchy"]["corpus_tier_assignment"]["legacy"]  # Director 승인 UNASSIGNED null을 읽는다.
    path = raw_root / str(entry["relative_path"])  # exact allowlist workbook path만 연다.
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)  # formula가 아닌 cached values를 bounded-memory로 읽는다.
    try:  # workbook handle을 오류 시에도 닫기 위해 보호 구간을 시작한다.
        assert len(workbook.sheetnames) == 1  # frozen allowlist workbook의 single-sheet 구조를 확인한다.
        sheet_name = workbook.sheetnames[0]  # raw sheet 이름을 provenance에 보존한다.
        worksheet = workbook[sheet_name]  # 첫이자 유일한 worksheet를 선택한다.
        rows = worksheet.iter_rows(values_only=True)  # cell 객체를 만들지 않고 scalar row만 streaming한다.
        header_values = next(rows)  # 첫 physical row를 header로 읽는다.
        headers = [str(value) if value is not None else None for value in header_values]  # blank trailing header를 null로 보존한다.
        assert "원문" in headers and "번역문" in headers  # paired raw text columns 존재를 확인한다.
        count = 0  # workbook별 실제 data rows를 센다.
        for physical_row_number, values in enumerate(rows, start=2):  # header 다음 physical row 번호 2부터 순회한다.
            record = {header: value for header, value in zip(headers, values, strict=True) if header is not None}  # blank header column을 제외하고 전체 raw row mapping을 만든다.
            ko_text = record.get("원문")  # Legacy paired KO raw field를 읽는다.
            en_text = record.get("번역문")  # Legacy paired EN raw field를 읽는다.
            assert isinstance(ko_text, str) and ko_text != ""  # structurally valid KO raw text를 확인한다.
            assert isinstance(en_text, str) and en_text != ""  # structurally valid EN raw text를 확인한다.
            locator_payload = {"physical_row_number": physical_row_number, "relative_path": str(entry["relative_path"]), "sheet_name": sheet_name}  # workbook/file+sheet+physical row locator를 구성한다.
            source_record_id = canonical_json_text(locator_payload)  # bare ID가 아닌 deterministic composite locator를 source record identity로 쓴다.
            native_id, native_sid, subdomain_raw, source_provenance_raw = _legacy_auxiliary(record, domain_raw)  # source-native metadata를 별도 보존한다.
            yield _base_staging_row(  # Legacy raw record를 공통 staging row로 만든다.
                source_id=source_id,
                source_tier=source_tier,
                domain=str(domain),
                direction_raw="UNKNOWN",
                direction_review=False,
                ko_text=ko_text,
                en_text=en_text,
                source_license_note="UNKNOWN — no license metadata in source",
                source_record_id=source_record_id,
                raw_locator=source_record_id,
                domain_raw=domain_raw,
                subdomain_raw=subdomain_raw,
                source_provenance_raw=source_provenance_raw,
                is_validation=False,
                mt_field_present=False,
                logical_corpus=logical_corpus,
                entry=entry,
                raw_sheet_name=sheet_name,
                raw_physical_row_number=physical_row_number,
                source_native_id=native_id,
                source_native_sid=native_sid,
                raw_metadata_json=_raw_metadata_without_text(record, {"원문", "번역문"}),
            )
            count += 1  # 성공적으로 만든 data row 수를 증가시킨다.
        assert count == int(entry["record_count"]), f"XLSX record count mismatch: {entry['relative_path']} expected={entry['record_count']} observed={count}"  # worksheet 종료 시 exact count를 확인한다.
    finally:  # 성공/실패와 무관하게 workbook resource를 해제한다.
        workbook.close()  # XLSX zip/file handle을 명시적으로 닫는다.


def write_staging_registry(
    raw_root: Path,
    entries: Sequence[Mapping[str, Any]],
    research: Mapping[str, Any],
    raw_manifest_sha256: str,
    staging_path: Path,
) -> dict[str, int]:
    """
    /**
     * @purpose exact allowlist의 565만 raw records를 bounded-memory staging Parquet으로 쓴다.
     * @spec_ref D-01 ingest contract; Notebook 01 input→ingest
     * @param raw_root immutable raw root
     * @param entries hash-verified canonical allowlist
     * @param research Claude-frozen research config
     * @param raw_manifest_sha256 source snapshot identity hash
     * @param staging_path project-contained interim Parquet path
     * @return family별 observed row count
     * @raises AssertionError schema/count 조건 불일치 시
     * @validation batch schema exact equality와 family total exact oracle로 검사한다.
     * @artifact data/interim/PAIR_REGISTRY_v001_staging.parquet
     */
    """
    staging_path.parent.mkdir(parents=True, exist_ok=True)  # project-contained ignored interim directory를 idempotent하게 만든다.
    temporary = staging_path.with_suffix(staging_path.suffix + ".partial")  # 불완전 staging을 canonical 경로와 분리한다.
    if temporary.exists():  # 이전 중단 실행의 partial file이 있으면 정확한 파일만 정리한다.
        temporary.unlink()  # 재실행 시 stale rows가 섞이지 않게 recoverable interim partial만 삭제한다.
    schema = staging_registry_schema()  # 모든 batch에 적용할 explicit nullable/type contract를 읽는다.
    writer = pq.ParquetWriter(temporary, schema=schema, compression="zstd", use_dictionary=True, write_statistics=True)  # deterministic column schema와 압축으로 writer를 연다.
    counts = {"025": 0, "026": 0, "LEGACY": 0}  # 실제 emitted rows를 family별 누적한다.
    batch: list[dict[str, Any]] = []  # bounded-memory batch rows를 누적한다.
    try:  # 예외 시에도 Parquet writer footer 처리/close를 보장한다.
        for entry in entries:  # contract 순서대로 canonical physical files를 정확히 한 번 처리한다.
            row_iterator = iter_json_rows(raw_root, entry, research, raw_manifest_sha256) if entry["format"] == "JSON" else iter_xlsx_rows(raw_root, entry, research, raw_manifest_sha256)  # format별 streaming reader만 선택한다.
            for row in row_iterator:  # source record를 한 행씩 소비한다.
                batch.append(row)  # bounded batch에 schema-ready mapping을 추가한다.
                counts[str(entry["logical_corpus"])] += 1  # family observed row count를 증가시킨다.
                if len(batch) >= BATCH_SIZE:  # batch가 고정 상한에 도달했는지 확인한다.
                    table = pa.Table.from_pylist(batch, schema=schema)  # explicit schema로 Arrow table을 만든다.
                    writer.write_table(table, row_group_size=BATCH_SIZE)  # 한 batch를 한 bounded row group으로 기록한다.
                    batch.clear()  # 기록된 Python row 객체를 즉시 해제한다.
        if batch:  # 마지막 partial batch가 남았는지 확인한다.
            table = pa.Table.from_pylist(batch, schema=schema)  # 남은 rows도 같은 schema로 변환한다.
            writer.write_table(table, row_group_size=BATCH_SIZE)  # 마지막 row group을 기록한다.
            batch.clear()  # 완료된 rows를 해제한다.
    finally:  # 성공/실패 여부와 무관하게 writer handle을 닫는다.
        writer.close()  # Parquet footer와 file handle을 확정한다.
    os.replace(temporary, staging_path)  # 완성된 staging만 atomic하게 최종 interim 경로로 승격한다.
    return counts  # notebook assertion과 source registry에 사용할 observed counts를 반환한다.


def _sql_path(path: Path) -> str:
    """
    /**
     * @purpose trusted project path를 DuckDB SQL string literal에 안전하게 넣는다.
     * @spec_ref engineering path containment
     * @param path SQL에서 참조할 local path
     * @return single-quote escaped absolute path
     * @raises 없음
     * @validation apostrophe 포함 path test로 검사한다.
     * @artifact 없음
     */
    """
    return str(path.resolve()).replace("'", "''")  # SQL literal delimiter와 충돌하는 apostrophe만 두 번 쓴다.


def resolve_duckdb_memory_limit(environ: Mapping[str, str] | None = None) -> str:
    """
    /**
     * @purpose DuckDB research-job memory cap을 안전한 기본값 또는 명시적 환경 override에서 결정한다.
     * @spec_ref INC-001 post-incident reproducibility engineering policy
     * @param environ 테스트 가능한 환경 mapping; None이면 process environment
     * @return 정규화된 DuckDB memory size 문자열
     * @raises ValueError override가 양의 정수 KB/MB/GB/TB 형식이 아닌 경우
     * @validation default=8GB, conservative override=6GB, invalid override fail-fast 단위 테스트
     * @artifact 없음
     */
    """
    source = os.environ if environ is None else environ  # production 환경과 주입된 test mapping을 명확히 분리한다.
    raw_value = source.get(DUCKDB_MEMORY_LIMIT_ENV, DEFAULT_DUCKDB_MEMORY_LIMIT)  # override가 없으면 안전한 8GB 기본값을 사용한다.
    match = _DUCKDB_MEMORY_LIMIT_PATTERN.fullmatch(raw_value.strip())  # 공백 외 추가 token이나 SQL 구문을 허용하지 않는다.
    if match is None:  # 지원하지 않는 값은 DuckDB 실행 전에 중단해야 한다.
        raise ValueError(f"{DUCKDB_MEMORY_LIMIT_ENV} must be a positive integer followed by KB, MB, GB, or TB; got {raw_value!r}")  # 오타와 unsafe override를 fail-fast한다.
    return f"{match.group('amount')}{match.group('unit').upper()}"  # case/공백 차이를 제거해 manifest-independent runtime setting을 만든다.


def finalize_registry(staging_path: Path, final_path: Path, runtime_dir: Path) -> None:
    """
    /**
     * @purpose duplicate group 전체의 대표·방향 resolution·conflict flags를 계산해 canonical registry를 만든다.
     * @spec_ref G1_PAIR_REGISTRY_PRECONTRACT_v1 §7; D-01 duplicate_representative
     * @param staging_path raw-record staging Parquet
     * @param final_path canonical PAIR_REGISTRY_v001 Parquet
     * @param runtime_dir project-contained DuckDB spill directory
     * @return None
     * @raises duckdb.Error SQL/group finalization 실패 시
     * @validation 대표=min(pair_id), mixed direction=UNKNOWN, conflict flags exact assertions로 검사한다.
     * @artifact data/registry/PAIR_REGISTRY_v001.parquet
     */
    """
    final_path.parent.mkdir(parents=True, exist_ok=True)  # canonical registry directory를 idempotent하게 만든다.
    runtime_dir.mkdir(parents=True, exist_ok=True)  # DuckDB spill을 project containment 안에 둔다.
    temporary = final_path.with_suffix(final_path.suffix + ".partial")  # 불완전 final artifact를 canonical 경로와 분리한다.
    if temporary.exists():  # 이전 중단 실행의 exact partial file이 남았는지 확인한다.
        temporary.unlink()  # stale partial만 제거하고 raw/canonical 입력은 건드리지 않는다.
    staging_literal = _sql_path(staging_path)  # staging absolute path를 SQL literal용으로 escape한다.
    output_literal = _sql_path(temporary)  # partial output absolute path를 SQL literal용으로 escape한다.
    runtime_literal = _sql_path(runtime_dir)  # spill directory absolute path를 SQL literal용으로 escape한다.
    memory_limit = resolve_duckdb_memory_limit()  # 8GB 기본값 또는 검증된 operator override를 실행 직전에 확정한다.
    group_derived = {  # 최종 schema에서 group resolution SQL 표현이 필요한 columns를 정의한다.
        "translation_direction": "CASE WHEN g.direction_value_count = 1 THEN g.single_direction ELSE 'UNKNOWN' END AS translation_direction",
        "representative_pair_id": "g.representative_pair_id AS representative_pair_id",
        "translation_direction_review_flag": "(s.translation_direction_review_flag OR g.direction_conflict_flag) AS translation_direction_review_flag",
        "direction_conflict_flag": "g.direction_conflict_flag AS direction_conflict_flag",
        "domain_conflict_flag": "g.domain_conflict_flag AS domain_conflict_flag",
        "source_id_conflict_flag": "g.source_id_conflict_flag AS source_id_conflict_flag",
        "source_provenance_raw_conflict_flag": "g.source_provenance_raw_conflict_flag AS source_provenance_raw_conflict_flag",
    }
    select_columns = ",\n                    ".join(group_derived.get(name, f"s.{name}") for name in pair_registry_schema().names)  # final schema 순서대로 raw 또는 group-derived 표현을 조립한다.
    query = f"""
        COPY (
            WITH group_stats AS (
                SELECT
                    duplicate_group_id,
                    min(pair_id) AS representative_pair_id,
                    count(DISTINCT translation_direction_raw) AS direction_value_count,
                    min(translation_direction_raw) AS single_direction,
                    count(DISTINCT translation_direction_raw) > 1 AS direction_conflict_flag,
                    count(DISTINCT domain) > 1 AS domain_conflict_flag,
                    count(DISTINCT source_id) > 1 AS source_id_conflict_flag,
                    count(DISTINCT source_provenance_raw) > 1 AS source_provenance_raw_conflict_flag
                FROM read_parquet('{staging_literal}')
                GROUP BY duplicate_group_id
            )
            SELECT
                    {select_columns}
            FROM read_parquet('{staging_literal}') AS s
            JOIN group_stats AS g USING (duplicate_group_id)
            ORDER BY s.pair_id
        ) TO '{output_literal}' (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 100000)
    """  # group semantics와 deterministic pair_id ordering을 하나의 out-of-core query로 고정한다.
    connection = duckdb.connect()  # 외부 DB state 없는 in-memory DuckDB connection을 만든다.
    try:  # 성공/실패와 무관하게 connection을 닫기 위한 보호 구간을 시작한다.
        connection.execute("SET preserve_insertion_order = false")  # 대용량 group/sort의 불필요한 insertion-order memory 비용을 줄인다.
        connection.execute(f"SET memory_limit = '{memory_limit}'")  # 검증된 runtime cap으로 host memory pressure를 제한한다.
        connection.execute(f"SET temp_directory = '{runtime_literal}'")  # project-contained 경로에만 spill하도록 제한한다.
        connection.execute(query)  # group-resolution과 canonical Parquet 생성을 실행한다.
    finally:  # query 결과와 무관하게 DuckDB resource를 해제한다.
        connection.close()  # file handles와 temporary execution state를 닫는다.
    os.replace(temporary, final_path)  # 완성된 final artifact만 canonical filename으로 atomic 승격한다.


def _scalar(connection: duckdb.DuckDBPyConnection, query: str, parameters: Sequence[Any] | None = None) -> Any:
    """
    /**
     * @purpose 단일 값 validation SQL의 결과를 명시적으로 추출한다.
     * @spec_ref fail-fast validation utility
     * @param connection active DuckDB connection
     * @param query exactly one row/one column SQL
     * @param parameters optional bound parameters
     * @return scalar result
     * @raises AssertionError 결과 shape가 1x1이 아닌 경우
     * @validation utility 단위 테스트로 검사한다.
     * @artifact reconciliation metrics
     */
    """
    row = connection.execute(query, parameters or []).fetchone()  # parameter binding으로 scalar query를 실행한다.
    assert row is not None and len(row) == 1  # validation query가 정확히 한 scalar를 반환하는지 확인한다.
    return row[0]  # scalar 값을 반환한다.


def validate_registry(final_path: Path, ingest_contract: Mapping[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """
    /**
     * @purpose D-01 row/key/hash/group semantics와 independent duplicate oracle을 실제 final Parquet에서 검증한다.
     * @spec_ref G1 Codex test oracle directive
     * @param final_path canonical PAIR_REGISTRY_v001 Parquet
     * @param ingest_contract data-recon independent expectations
     * @return machine metrics와 human reconciliation rows
     * @raises AssertionError 어떤 PASS 기준도 불일치하는 경우
     * @validation exact counts/oracles가 곧 이 함수의 fail-fast assertions다.
     * @artifact PAIR_REGISTRY_RECONCILIATION_v001.csv, manifest.validation
     */
    """
    path_literal = _sql_path(final_path)  # canonical Parquet path를 SQL용으로 escape한다.
    relation = f"read_parquet('{path_literal}')"  # 반복 validation query의 source relation을 고정한다.
    expected_counts = ingest_contract["expected_logical_record_counts"]  # family/total exact row oracle을 읽는다.
    duplicate_oracle = ingest_contract["duplicate_baseline_oracle"]  # independent duplicate metrics를 읽는다.
    connection = duckdb.connect()  # canonical Parquet만 읽는 in-memory validation connection을 만든다.
    try:  # 모든 query 후 connection을 확실히 닫기 위한 보호 구간을 시작한다.
        total_rows = int(_scalar(connection, f"SELECT count(*) FROM {relation}"))  # 전체 logical row 수를 계산한다.
        distinct_pair_ids = int(_scalar(connection, f"SELECT count(DISTINCT pair_id) FROM {relation}"))  # primary key distinct 수를 계산한다.
        null_pair_ids = int(_scalar(connection, f"SELECT count(*) FROM {relation} WHERE pair_id IS NULL"))  # primary key null 수를 계산한다.
        empty_raw_pairs = int(_scalar(connection, f"SELECT count(*) FROM {relation} WHERE ko_text_raw IS NULL OR en_text_raw IS NULL OR ko_text_raw = '' OR en_text_raw = ''"))  # structural raw pair 누락을 계산한다.
        phase1_status_violations = int(
            _scalar(connection, f"SELECT count(*) FROM {relation} WHERE pair_quality_status <> 'review' OR pair_quality_score IS NOT NULL OR normalization_status <> 'NOT_GENERATED_PHASE1' OR qc_stage_status <> 'PENDING_PHASE2'")
        )  # Phase 2 semantics가 조기 적용됐는지 검사한다.
        normalization_value_violations = int(_scalar(connection, f"SELECT count(*) FROM {relation} WHERE ko_text_nfc IS NOT NULL OR en_text_nfc IS NOT NULL OR ko_text_analysis IS NOT NULL OR en_text_analysis IS NOT NULL"))  # normalization 금지 위반을 계산한다.
        representative_violations = int(
            _scalar(connection, f"SELECT count(*) FROM {relation} r JOIN (SELECT duplicate_group_id, min(pair_id) AS expected_rep FROM {relation} GROUP BY duplicate_group_id) g USING (duplicate_group_id) WHERE r.representative_pair_id <> g.expected_rep")
        )  # 대표=min(pair_id) 규칙 위반을 계산한다.
        direction_resolution_violations = int(
            _scalar(
                connection,
                f"SELECT count(*) FROM {relation} r "
                f"JOIN (SELECT duplicate_group_id, count(DISTINCT translation_direction_raw) AS n_dir, min(translation_direction_raw) AS only_dir FROM {relation} GROUP BY duplicate_group_id) g USING (duplicate_group_id) "
                "WHERE r.translation_direction <> CASE WHEN g.n_dir = 1 THEN g.only_dir ELSE 'UNKNOWN' END OR r.direction_conflict_flag <> (g.n_dir > 1)",
            )
        )  # group-resolved 방향과 conflict flag 위반을 계산한다.
        raw_hash_linkage_violations = int(_scalar(connection, f"SELECT count(*) FROM {relation} WHERE raw_file_relative_path IS NULL OR raw_file_sha256 IS NULL OR length(raw_file_sha256) <> 64"))  # 모든 row가 verified input hash와 연결되는지 검사한다.
        row_counts = {str(corpus): int(count) for corpus, count in connection.execute(f"SELECT logical_corpus, count(*) FROM {relation} GROUP BY logical_corpus ORDER BY logical_corpus").fetchall()}  # family별 actual rows를 계산한다.
        duplicate_metrics = {  # corpus 내부 duplicate after-first와 group 수를 계산한다.
            str(corpus): {"duplicate_pair_rows_after_first_occurrence": int(after_first), "distinct_duplicate_pair_groups": int(groups)}
            for corpus, after_first, groups in connection.execute(
                f"SELECT logical_corpus, sum(group_count - 1), count(*) FROM (SELECT logical_corpus, duplicate_group_id, count(*) AS group_count FROM {relation} GROUP BY logical_corpus, duplicate_group_id HAVING count(*) > 1) GROUP BY logical_corpus ORDER BY logical_corpus"
            ).fetchall()
        }
        cross_direction_025 = int(_scalar(connection, f"SELECT count(*) FROM (SELECT duplicate_group_id FROM {relation} WHERE logical_corpus='025' GROUP BY duplicate_group_id HAVING count(DISTINCT translation_direction_raw) > 1)"))  # 025 mixed-direction distinct content groups를 계산한다.
        cross_pairs = {  # corpus pair별 shared distinct content group 수를 계산한다.
            "025_vs_026": int(_scalar(connection, f"SELECT count(*) FROM (SELECT duplicate_group_id FROM {relation} WHERE logical_corpus IN ('025','026') GROUP BY duplicate_group_id HAVING count(DISTINCT logical_corpus)=2)")),
            "025_vs_legacy": int(_scalar(connection, f"SELECT count(*) FROM (SELECT duplicate_group_id FROM {relation} WHERE logical_corpus IN ('025','LEGACY') GROUP BY duplicate_group_id HAVING count(DISTINCT logical_corpus)=2)")),
            "026_vs_legacy": int(_scalar(connection, f"SELECT count(*) FROM (SELECT duplicate_group_id FROM {relation} WHERE logical_corpus IN ('026','LEGACY') GROUP BY duplicate_group_id HAVING count(DISTINCT logical_corpus)=2)")),
        }
        legacy_news2_culture = int(
            _scalar(connection, f"SELECT count(*) FROM (SELECT duplicate_group_id FROM {relation} WHERE canonical_ingest_role IN ('LEGACY_3_문어체_뉴스(2)','LEGACY_4_문어체_한국문화') GROUP BY duplicate_group_id HAVING count(DISTINCT canonical_ingest_role)=2)")
        )  # Legacy News2↔Culture anomaly oracle를 계산한다.
        sn_collisions = {  # JSON family에서 canonical sn/source_record_id collision 수를 계산한다.
            corpus: int(_scalar(connection, f"SELECT count(*) FROM (SELECT source_record_id FROM {relation} WHERE logical_corpus=? GROUP BY source_record_id HAVING count(*)>1)", [corpus])) for corpus in ("025", "026")
        }
        actual_file_count = int(_scalar(connection, f"SELECT count(DISTINCT raw_file_relative_path) FROM {relation}"))  # canonical ingested physical file 수를 계산한다.
    finally:  # validation 성공/실패와 무관하게 connection을 닫는다.
        connection.close()  # query state와 file handles를 해제한다.
    assertions = {  # 각 acceptance criterion의 observed/expected/pass 값을 표준화한다.
        "total_rows": (total_rows, int(expected_counts["total"])),
        "pair_id_unique": (distinct_pair_ids, total_rows),
        "pair_id_null": (null_pair_ids, 0),
        "raw_pair_structural_missing": (empty_raw_pairs, 0),
        "phase1_status_violations": (phase1_status_violations, 0),
        "normalization_value_violations": (normalization_value_violations, 0),
        "representative_rule_violations": (representative_violations, 0),
        "direction_resolution_violations": (direction_resolution_violations, 0),
        "raw_hash_linkage_violations": (raw_hash_linkage_violations, 0),
        "canonical_physical_file_count": (actual_file_count, 16),
        "025_rows": (row_counts.get("025"), int(expected_counts["025"])),
        "026_rows": (row_counts.get("026"), int(expected_counts["026"])),
        "LEGACY_rows": (row_counts.get("LEGACY"), int(expected_counts["LEGACY"])),
        "025_duplicate_after_first": (duplicate_metrics["025"]["duplicate_pair_rows_after_first_occurrence"], int(duplicate_oracle["025"]["duplicate_pair_rows_after_first_occurrence"])),
        "025_duplicate_groups": (duplicate_metrics["025"]["distinct_duplicate_pair_groups"], int(duplicate_oracle["025"]["distinct_duplicate_pair_groups"])),
        "026_duplicate_after_first": (duplicate_metrics["026"]["duplicate_pair_rows_after_first_occurrence"], int(duplicate_oracle["026"]["duplicate_pair_rows_after_first_occurrence"])),
        "026_duplicate_groups": (duplicate_metrics["026"]["distinct_duplicate_pair_groups"], int(duplicate_oracle["026"]["distinct_duplicate_pair_groups"])),
        "025_cross_direction_distinct": (cross_direction_025, int(duplicate_oracle["cross_direction_distinct_shared_pair_digests"])),
        "025_vs_026": (cross_pairs["025_vs_026"], int(duplicate_oracle["cross_corpus_raw_exact_pair_overlap"]["025_vs_026"])),
        "025_vs_legacy": (cross_pairs["025_vs_legacy"], int(duplicate_oracle["cross_corpus_raw_exact_pair_overlap"]["025_vs_legacy"])),
        "026_vs_legacy": (cross_pairs["026_vs_legacy"], int(duplicate_oracle["cross_corpus_raw_exact_pair_overlap"]["026_vs_legacy"])),
        "Legacy_News2_vs_Culture": (legacy_news2_culture, int(duplicate_oracle["legacy_news2_culture_anomaly"]["shared_distinct_pair_digest_count"])),
        "025_sn_collision_groups": (sn_collisions["025"], 0),
        "026_sn_collision_groups": (sn_collisions["026"], 0),
        "accidental_alias_double_ingest": (actual_file_count - 16, 0),
    }
    reconciliation_rows = [  # human-readable CSV grain을 assertion 1개당 1행으로 구성한다.
        {"check_id": check_id, "observed": observed, "expected": expected, "status": "PASS" if observed == expected else "FAIL"} for check_id, (observed, expected) in assertions.items()
    ]
    failures = [row for row in reconciliation_rows if row["status"] != "PASS"]  # 불일치 assertion rows만 모은다.
    assert not failures, f"G1 registry oracle mismatch: {failures}"  # oracle를 조용히 갱신하지 않고 즉시 중단한다.
    metrics = {  # manifest에 저장할 machine-readable validation summary를 구성한다.
        "cross_corpus_overlap": cross_pairs,
        "cross_direction_025": cross_direction_025,
        "duplicate_metrics": duplicate_metrics,
        "legacy_news2_culture": legacy_news2_culture,
        "pair_id_distinct": distinct_pair_ids,
        "row_counts": {**row_counts, "total": total_rows},
        "sn_collision_groups": sn_collisions,
        "status": "PASS",
    }
    return metrics, reconciliation_rows  # artifact persistence에 필요한 machine/human 결과를 반환한다.


def write_reconciliation_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    """
    /**
     * @purpose validation assertion 결과를 human-readable UTF-8 CSV로 원자 저장한다.
     * @spec_ref D-01 outputs contract
     * @param path PAIR_REGISTRY_RECONCILIATION_v001.csv 경로
     * @param rows check_id/observed/expected/status 행 목록
     * @return None
     * @raises OSError 또는 ValueError 저장/column 계약 위반 시
     * @validation 재독 후 row 수와 모든 status=PASS를 검사한다.
     * @artifact outputs/reports/PAIR_REGISTRY_RECONCILIATION_v001.csv
     */
    """
    path.parent.mkdir(parents=True, exist_ok=True)  # 승인된 reports directory를 idempotent하게 만든다.
    temporary = path.with_suffix(path.suffix + ".tmp")  # incomplete CSV를 final 경로와 분리한다.
    fieldnames = ["check_id", "observed", "expected", "status"]  # stable human table column 순서를 고정한다.
    with temporary.open("w", encoding="utf-8", newline="") as stream:  # UTF-8과 CSV newline 규칙으로 임시 파일을 연다.
        writer = csv.DictWriter(stream, fieldnames=fieldnames)  # 명시적 column 순서의 writer를 만든다.
        writer.writeheader()  # human-readable header를 기록한다.
        for row in rows:  # validation check를 계약 순서대로 기록한다.
            writer.writerow({name: row[name] for name in fieldnames})  # 승인 columns만 정확히 한 행으로 저장한다.
    os.replace(temporary, path)  # 완성된 CSV만 final report 경로로 atomic 승격한다.


def write_source_registry(
    path: Path,
    entries: Sequence[Mapping[str, Any]],
    research: Mapping[str, Any],
    raw_manifest_sha256: str,
    row_counts: Mapping[str, int],
) -> None:
    """
    /**
     * @purpose 025/026/Legacy family-level provenance와 ingest summary를 source registry로 저장한다.
     * @spec_ref SSOT §9.3, §12.1; G1_PAIR_REGISTRY_PRECONTRACT_v1 §1-2
     * @param path SOURCE_REGISTRY_v001 Parquet 경로
     * @param entries verified canonical ingest allowlist
     * @param research Claude-frozen research config
     * @param raw_manifest_sha256 verified raw snapshot hash
     * @param row_counts final pair registry family counts
     * @return None
     * @raises AssertionError 3개 source 또는 count 계약 불일치 시
     * @validation source_id uniqueness, row count=3, expected=observed를 검사한다.
     * @artifact data/registry/SOURCE_REGISTRY_v001.parquet
     */
    """
    path.parent.mkdir(parents=True, exist_ok=True)  # canonical registry directory를 idempotent하게 만든다.
    expected_counts = {"025": 2_700_345, "026": 1_350_162, "LEGACY": 1_602_418}  # frozen D-01 logical family counts를 명시한다.
    rows: list[dict[str, Any]] = []  # family grain source rows를 누적한다.
    for corpus in ("025", "026", "LEGACY"):  # stable family order로 source rows를 만든다.
        corpus_entries = [entry for entry in entries if entry["logical_corpus"] == corpus]  # 해당 family canonical files만 선택한다.
        config_key = "legacy" if corpus == "LEGACY" else corpus  # research config의 lowercase legacy key 차이를 명시적으로 처리한다.
        portfolio = research["source_portfolio"][config_key]  # source role/eligibility/provenance closure를 authoritative config에서 읽는다.
        tier = research["source_hierarchy"]["corpus_tier_assignment"][config_key]  # Director 승인 Tier 또는 null을 읽는다.
        license_note = "UNKNOWN — no license metadata in source" if corpus == "LEGACY" else "Raw records self-assert license=open; redistribution permission NOT VERIFIED"  # family-level evidence boundary를 명시한다.
        input_files = [{"relative_path": str(entry["relative_path"]), "sha256": str(entry["sha256"]), "record_count": int(entry["record_count"])} for entry in corpus_entries]  # canonical file provenance를 compact inventory로 만든다.
        rows.append(  # one family/acquisition snapshot을 한 source row로 추가한다.
            {
                "source_id": source_id_for(corpus, raw_manifest_sha256),
                "logical_corpus": corpus,
                "source_tier": tier,
                "research_role": str(portfolio["research_role"]),
                "primary_analysis_eligible": bool(portfolio["primary_analysis_eligible"]),
                "provenance_closure_status": str(portfolio["provenance_closure_status"]),
                "official_dataset_id": None,
                "raw_manifest_sha256": raw_manifest_sha256,
                "source_license_note": license_note,
                "expected_row_count": expected_counts[corpus],
                "observed_row_count": int(row_counts[corpus]),
                "input_file_count": len(corpus_entries),
                "input_files_json": canonical_json_text(input_files),
                "pair_version": PAIR_VERSION,
            }
        )
    assert len(rows) == 3 and len({row["source_id"] for row in rows}) == 3  # source registry grain과 primary key uniqueness를 확인한다.
    assert all(row["expected_row_count"] == row["observed_row_count"] for row in rows)  # source family row counts가 exact인지 확인한다.
    table = pa.Table.from_pylist(rows, schema=source_registry_schema())  # explicit nullable/type schema로 Arrow table을 만든다.
    temporary = path.with_suffix(path.suffix + ".tmp")  # 불완전 source registry를 final 경로와 분리한다.
    pq.write_table(table, temporary, compression="zstd", use_dictionary=True, write_statistics=True)  # small canonical source table을 Parquet으로 저장한다.
    os.replace(temporary, path)  # 완성된 source registry만 canonical 경로로 atomic 승격한다.


def git_head(project_root: Path) -> str:
    """
    /**
     * @purpose 실행 코드가 속한 Git HEAD SHA를 artifact manifest에 연결한다.
     * @spec_ref SSOT §30.2, artifact contract
     * @param project_root active Git worktree root
     * @return 40자리 commit SHA
     * @raises subprocess.CalledProcessError Git 조회 실패 시
     * @validation git rev-parse HEAD CLI 출력과 exact equality로 검사한다.
     * @artifact manifest.code_commit
     */
    """
    return subprocess.run(["git", "rev-parse", "HEAD"], cwd=project_root, check=True, capture_output=True, text=True).stdout.strip()  # shell interpolation 없이 현재 worktree HEAD를 읽는다.


def schema_fingerprint(schema: pa.Schema) -> str:
    """
    /**
     * @purpose Arrow schema의 ordered field/name/type/nullability를 canonical SHA-256으로 고정한다.
     * @spec_ref artifact schema/hash contract
     * @param schema fingerprint 대상 PyArrow schema
     * @return canonical schema SHA-256
     * @raises 없음
     * @validation field order/type/nullability 변화가 hash를 바꾸는지 검사한다.
     * @artifact manifest.schema_sha256
     */
    """
    payload = [{"name": field.name, "nullable": field.nullable, "type": str(field.type)} for field in schema]  # schema semantics를 JSON-compatible ordered rows로 만든다.
    return hashlib.sha256(canonical_json_bytes(payload)).hexdigest()  # project canonical JSON bytes의 SHA-256을 반환한다.


def build_manifest(
    *,
    execution_root: Path,
    canonical_root: Path,
    raw_root: Path,
    research_config_path: Path,
    ingest_contract_path: Path,
    notebook_path: Path,
    pair_registry_path: Path,
    source_registry_path: Path,
    reconciliation_path: Path,
    entries: Sequence[Mapping[str, Any]],
    raw_manifest_sha256: str,
    metrics: Mapping[str, Any],
) -> dict[str, Any]:
    """
    /**
     * @purpose D-01 artifacts, contracts, inputs, code commit, schema, row counts를 하나의 provenance manifest로 묶는다.
     * @spec_ref SSOT §30.2, §38; D-01 output directive
     * @param keyword-only execution root, canonical root, paths, contracts, metrics evidence
     * @return JSON-serializable PAIR_REGISTRY_MANIFEST_v001 payload
     * @raises FileNotFoundError required artifact/contract가 없는 경우
     * @validation 저장 후 artifact hashes, row counts, canonical_root containment를 재검사한다.
     * @artifact outputs/manifests/PAIR_REGISTRY_MANIFEST_v001.json
     */
    """
    input_files = [  # physical input provenance를 contract 순서대로 manifest rows로 만든다.
        {"logical_corpus": str(entry["logical_corpus"]), "record_count": int(entry["record_count"]), "relative_path": str(entry["relative_path"]), "sha256": str(entry["sha256"])} for entry in entries
    ]
    return {  # release-critical provenance와 validation을 한 payload로 반환한다.
        "artifact_id": "PAIR_REGISTRY_MANIFEST_v001",
        "canonical_root": str(canonical_root.resolve()),
        "code_commit": git_head(execution_root),
        "config_sha256": sha256_file(research_config_path),
        "created_at": dt.datetime.now(dt.UTC).astimezone().isoformat(),
        "determinism_status": "DETERMINISTIC_GIVEN_IDENTICAL_INPUT_BYTES_CONTRACTS_CODE_AND_PACKAGE_LOCK",
        "ingest_contract_sha256": sha256_file(ingest_contract_path),
        "input_file_hashes": input_files,
        "notebook_input_sha256": sha256_file(notebook_path),
        "pair_registry": {
            "path": pair_registry_path.relative_to(execution_root).as_posix(),
            "row_count": int(metrics["row_counts"]["total"]),
            "schema_sha256": schema_fingerprint(pair_registry_schema()),
            "schema_version": PAIR_REGISTRY_SCHEMA_VERSION,
            "sha256": sha256_file(pair_registry_path),
        },
        "pair_version": PAIR_VERSION,
        "raw_root": str(raw_root.resolve()),
        "raw_source_manifest_sha256": raw_manifest_sha256,
        "reconciliation": {
            "path": reconciliation_path.relative_to(execution_root).as_posix(),
            "row_count": sum(1 for _ in reconciliation_path.open(encoding="utf-8")) - 1,
            "sha256": sha256_file(reconciliation_path),
        },
        "research_contract_path": str(research_config_path.resolve()),
        "source_registry": {
            "path": source_registry_path.relative_to(execution_root).as_posix(),
            "row_count": 3,
            "schema_sha256": schema_fingerprint(source_registry_schema()),
            "schema_version": SOURCE_REGISTRY_SCHEMA_VERSION,
            "sha256": sha256_file(source_registry_path),
        },
        "validation": dict(metrics),
    }


def persist_manifest(path: Path, payload: Mapping[str, Any]) -> None:
    """
    /**
     * @purpose pair registry provenance manifest를 project canonical JSON 규칙으로 저장하고 roundtrip 검증한다.
     * @spec_ref SSOT §30.2, §38; artifact contract
     * @param path PAIR_REGISTRY_MANIFEST_v001.json 경로
     * @param payload build_manifest 결과
     * @return None
     * @raises AssertionError 저장 roundtrip 또는 canonical root leak 검사 실패 시
     * @validation bytes 재파싱 equality와 worktree-path 금지를 검사한다.
     * @artifact outputs/manifests/PAIR_REGISTRY_MANIFEST_v001.json
     */
    """
    write_json(path, dict(payload))  # 공통 atomic UTF-8 JSON writer로 manifest를 저장한다.
    loaded = json.loads(path.read_text(encoding="utf-8"))  # 저장 artifact를 UTF-8로 다시 읽는다.
    assert loaded == payload  # serialization roundtrip에서 provenance 값이 바뀌지 않았는지 확인한다.
    assert ".agent_worktrees/codex" not in loaded["canonical_root"]  # canonical root field가 agent worktree를 가리키면 fail한다.
